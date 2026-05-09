"""Phase 7B — Advanced Reports service.

Provides:
- Vendor Performance Scorecard (on-time%, price stability, defect rate, lead time avg, total spend)
- Report Builder (lite): pick dimensions × metrics × filters → aggregated rows
- Pivot Matrix (Outlet × Brand × Category, etc.)
- Comparatives (MoM / YoY toggle for sales/expense/AP)
- Saved Reports CRUD (per user-owned definitions)
"""
from __future__ import annotations

import logging
import math
import statistics
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Optional

from core.db import get_db, serialize
from core.exceptions import ConflictError, NotFoundError, ValidationError

logger = logging.getLogger("aurora.reports")


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _parse_date(s: Optional[str], *, fallback_today: bool = False) -> Optional[datetime]:
    if not s:
        if fallback_today:
            return datetime.now(timezone.utc)
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError as e:
        raise ValidationError(f"Invalid date format (expect YYYY-MM-DD): {s}") from e


# ============================================================
# 1. VENDOR PERFORMANCE SCORECARD
# ============================================================
async def vendor_scorecard(
    *,
    vendor_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    top: int = 20,
) -> dict[str, Any]:
    """Compute vendor performance metrics from purchase_orders + goods_receipts.

    Metrics per vendor:
      - po_count, gr_count
      - total_spend (sum GR grand_total)
      - on_time_pct: GR receive_date <= PO expected_delivery_date
      - avg_lead_time_days: receive_date - sent_at (POs that have GR)
      - price_stability: 1 - (stddev(unit_cost) / avg(unit_cost)) per item, averaged
      - defect_rate: sum(qty_ordered - qty_received) / sum(qty_ordered)
      - late_delivery_count

    If vendor_id given → returns single vendor detail with per-PO breakdown.
    Else → returns list ranked by total_spend.
    """
    db = get_db()

    # Build common date filter on PO order_date
    date_filter: dict[str, Any] = {}
    if date_from:
        date_filter["$gte"] = date_from
    if date_to:
        date_filter["$lte"] = date_to

    po_match: dict[str, Any] = {"deleted_at": None}
    if date_filter:
        po_match["order_date"] = date_filter
    if vendor_id:
        po_match["vendor_id"] = vendor_id

    # Load all relevant POs
    pos: list[dict] = []
    async for po in db.purchase_orders.find(po_match):
        pos.append(po)

    if not pos:
        return {
            "vendors": [],
            "filters": {"vendor_id": vendor_id, "date_from": date_from, "date_to": date_to},
        }

    # Load all related GRs in one go
    po_ids = [p["id"] for p in pos]
    grs_by_po: dict[str, list[dict]] = {}
    async for gr in db.goods_receipts.find({"deleted_at": None, "po_id": {"$in": po_ids}}):
        grs_by_po.setdefault(gr["po_id"], []).append(gr)

    # Load vendor names
    vendor_ids = list({p["vendor_id"] for p in pos if p.get("vendor_id")})
    vendors_by_id: dict[str, dict] = {}
    async for v in db.vendors.find({"id": {"$in": vendor_ids}}):
        vendors_by_id[v["id"]] = {"id": v["id"], "name": v.get("name", v["id"]), "code": v.get("code", "")}

    # Group POs by vendor and compute metrics
    by_vendor: dict[str, dict[str, Any]] = {}
    for po in pos:
        vid = po["vendor_id"]
        bucket = by_vendor.setdefault(vid, {
            "vendor_id": vid,
            "vendor_name": vendors_by_id.get(vid, {}).get("name", vid),
            "vendor_code": vendors_by_id.get(vid, {}).get("code", ""),
            "po_count": 0,
            "gr_count": 0,
            "total_spend": 0.0,
            "on_time_count": 0,
            "late_count": 0,
            "lead_times": [],
            "qty_ordered": 0.0,
            "qty_received": 0.0,
            "item_costs": {},  # item_id -> [unit_cost,...]
            "po_breakdown": [],
        })
        bucket["po_count"] += 1
        grs = grs_by_po.get(po["id"], [])
        bucket["gr_count"] += len(grs)

        po_qty_ordered = sum(float(ln.get("qty", 0) or 0) for ln in po.get("lines", []))
        bucket["qty_ordered"] += po_qty_ordered

        # Track per-item unit cost from PO lines (price stability sample)
        for ln in po.get("lines", []):
            iid = ln.get("item_id")
            uc = float(ln.get("unit_cost", 0) or 0)
            if iid and uc > 0:
                bucket["item_costs"].setdefault(iid, []).append(uc)

        po_qty_received = 0.0
        po_spend = 0.0
        on_time_for_po = None  # tri-state: True/False/None
        for gr in grs:
            po_spend += float(gr.get("grand_total", 0) or 0)
            for gln in gr.get("lines", []):
                po_qty_received += float(gln.get("qty_received", 0) or 0)
            # Lead time
            try:
                if po.get("sent_at") and gr.get("receive_date"):
                    sent = datetime.fromisoformat(po["sent_at"].replace("Z", "+00:00"))
                    recv = datetime.strptime(gr["receive_date"], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    lead = (recv - sent).days
                    if lead >= 0:
                        bucket["lead_times"].append(lead)
            except Exception:  # noqa: BLE001
                pass
            # On-time check
            try:
                if po.get("expected_delivery_date") and gr.get("receive_date"):
                    exp_d = datetime.strptime(po["expected_delivery_date"], "%Y-%m-%d").date()
                    rcv_d = datetime.strptime(gr["receive_date"], "%Y-%m-%d").date()
                    on_time_for_po = rcv_d <= exp_d
            except Exception:  # noqa: BLE001
                pass

        if on_time_for_po is True:
            bucket["on_time_count"] += 1
        elif on_time_for_po is False:
            bucket["late_count"] += 1

        bucket["qty_received"] += po_qty_received
        bucket["total_spend"] += po_spend

        if vendor_id:  # single-vendor detail mode
            bucket["po_breakdown"].append({
                "po_id": po["id"],
                "doc_no": po.get("doc_no"),
                "order_date": po.get("order_date"),
                "expected_delivery_date": po.get("expected_delivery_date"),
                "grand_total": float(po.get("grand_total", 0) or 0),
                "status": po.get("status"),
                "gr_count": len(grs),
                "received_total": po_spend,
                "qty_ordered": po_qty_ordered,
                "qty_received": po_qty_received,
                "on_time": on_time_for_po,
                "doc_first_gr": grs[0].get("doc_no") if grs else None,
                "first_gr_id": grs[0].get("id") if grs else None,
            })

    # Compute derived metrics
    out_rows: list[dict[str, Any]] = []
    for vid, b in by_vendor.items():
        rated_pos = b["on_time_count"] + b["late_count"]
        on_time_pct = round(b["on_time_count"] / rated_pos * 100, 2) if rated_pos else None
        avg_lead = round(statistics.mean(b["lead_times"]), 2) if b["lead_times"] else None
        # Price stability: 1 - mean(stddev/mean) across items with >= 2 samples
        ratios: list[float] = []
        for iid, costs in b["item_costs"].items():
            if len(costs) >= 2 and statistics.mean(costs) > 0:
                ratio = statistics.stdev(costs) / statistics.mean(costs)
                ratios.append(ratio)
        price_stability = round((1.0 - statistics.mean(ratios)) * 100, 2) if ratios else None
        if price_stability is not None and price_stability < 0:
            price_stability = 0.0
        defect_rate = round(
            (b["qty_ordered"] - b["qty_received"]) / b["qty_ordered"] * 100, 2,
        ) if b["qty_ordered"] > 0 else None

        # Composite score 0-100 (weighted)
        components: list[float] = []
        if on_time_pct is not None:
            components.append(on_time_pct * 0.40)
        if price_stability is not None:
            components.append(price_stability * 0.25)
        if defect_rate is not None:
            components.append(max(0, 100 - defect_rate) * 0.20)
        if avg_lead is not None:
            # 0-day = 100, >=14 days = 0
            lead_score = max(0, 100 - (avg_lead / 14) * 100)
            components.append(lead_score * 0.15)
        composite = round(sum(components), 2) if components else None

        row: dict[str, Any] = {
            "vendor_id": vid,
            "vendor_name": b["vendor_name"],
            "vendor_code": b["vendor_code"],
            "po_count": b["po_count"],
            "gr_count": b["gr_count"],
            "total_spend": round(b["total_spend"], 2),
            "on_time_pct": on_time_pct,
            "late_count": b["late_count"],
            "avg_lead_time_days": avg_lead,
            "price_stability_pct": price_stability,
            "defect_rate_pct": defect_rate,
            "composite_score": composite,
        }
        if vendor_id:
            row["po_breakdown"] = sorted(
                b["po_breakdown"], key=lambda r: r.get("order_date") or "", reverse=True,
            )
            row["item_price_samples"] = [
                {"item_id": iid, "samples": costs, "mean": round(statistics.mean(costs), 2),
                 "stdev": round(statistics.stdev(costs), 2) if len(costs) >= 2 else 0,
                 "n": len(costs)}
                for iid, costs in b["item_costs"].items() if costs
            ][:30]
        out_rows.append(row)

    out_rows.sort(key=lambda r: r["total_spend"], reverse=True)
    if not vendor_id:
        out_rows = out_rows[:top]

    return {
        "vendors": out_rows,
        "filters": {"vendor_id": vendor_id, "date_from": date_from, "date_to": date_to, "top": top},
        "summary": {
            "vendor_count": len(out_rows),
            "total_spend": round(sum(r["total_spend"] for r in out_rows), 2),
            "avg_on_time_pct": round(
                statistics.mean([r["on_time_pct"] for r in out_rows if r["on_time_pct"] is not None]), 2,
            ) if [r for r in out_rows if r["on_time_pct"] is not None] else None,
        },
    }


# ============================================================
# 2. REPORT BUILDER (lite)
# ============================================================
SUPPORTED_DIMENSIONS: dict[str, str] = {
    # name -> source key in lookup map
    "outlet": "outlet_id",
    "brand": "brand_id",
    "vendor": "vendor_id",
    "category": "category_id",
    "month": "month",  # derived from date
}

SUPPORTED_METRICS: dict[str, str] = {
    "sales": "Sales (validated daily_sales grand_total)",
    "transaction_count": "Transaction count (validated daily_sales)",
    "cogs": "COGS (JE postings to COGS account)",
    "gross_profit": "Gross Profit (sales - cogs)",
    "ap_exposure": "AP Exposure (open GR grand_total)",
    "po_count": "PO count",
    "gr_count": "GR count",
    "purchase_value": "Purchase value (GR grand_total)",
}


async def report_builder(
    *,
    dimensions: list[str],
    metrics: list[str],
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
    outlet_ids: Optional[list[str]] = None,
    brand_ids: Optional[list[str]] = None,
    vendor_ids: Optional[list[str]] = None,
    category_ids: Optional[list[str]] = None,
    sort_by: Optional[str] = None,
    sort_dir: str = "desc",
    limit: int = 100,
) -> dict[str, Any]:
    """Run an ad-hoc aggregation. Returns rows keyed by dimension tuple + metric values."""
    if not dimensions:
        raise ValidationError("Pilih minimal 1 dimensi")
    if not metrics:
        raise ValidationError("Pilih minimal 1 metrik")
    bad_dims = [d for d in dimensions if d not in SUPPORTED_DIMENSIONS]
    if bad_dims:
        raise ValidationError(f"Dimensi tidak didukung: {bad_dims}")
    bad_metrics = [m for m in metrics if m not in SUPPORTED_METRICS]
    if bad_metrics:
        raise ValidationError(f"Metrik tidak didukung: {bad_metrics}")

    db = get_db()

    # Build a unified row keyed by dimension tuple from each source collection,
    # then compute metrics per unique combination.

    # Lookup maps for human-readable labels
    outlets_map: dict[str, dict] = {}
    brands_map: dict[str, dict] = {}
    vendors_map: dict[str, dict] = {}
    categories_map: dict[str, dict] = {}
    items_map: dict[str, dict] = {}

    if "outlet" in dimensions or outlet_ids:
        async for o in db.outlets.find({"deleted_at": None}):
            outlets_map[o["id"]] = {"id": o["id"], "name": o.get("name", o["id"]),
                                     "brand_id": o.get("brand_id"), "code": o.get("code", "")}
    if "brand" in dimensions or brand_ids:
        async for b in db.brands.find({"deleted_at": None}):
            brands_map[b["id"]] = {"id": b["id"], "name": b.get("name", b["id"]), "code": b.get("code", "")}
    if "vendor" in dimensions or vendor_ids:
        async for v in db.vendors.find({"deleted_at": None}):
            vendors_map[v["id"]] = {"id": v["id"], "name": v.get("name", v["id"]), "code": v.get("code", "")}
    if "category" in dimensions or category_ids:
        async for c in db.categories.find({"deleted_at": None}):
            categories_map[c["id"]] = {"id": c["id"], "name": c.get("name", c["id"]), "code": c.get("code", "")}
        async for it in db.items.find({"deleted_at": None}, {"id": 1, "category_id": 1}):
            items_map[it["id"]] = {"category_id": it.get("category_id")}

    # Build aggregation result per dim_tuple
    agg: dict[tuple, dict[str, float]] = {}

    # ---- Sales / transaction_count / gross_profit (sales side) ----
    if any(m in metrics for m in ("sales", "transaction_count", "gross_profit")):
        sales_match: dict = {"deleted_at": None, "status": "validated"}
        if period_from or period_to:
            sales_match["sales_date"] = {}
            if period_from:
                sales_match["sales_date"]["$gte"] = period_from
            if period_to:
                sales_match["sales_date"]["$lte"] = period_to
        if outlet_ids:
            sales_match["outlet_id"] = {"$in": outlet_ids}
        if brand_ids:
            sales_match["brand_id"] = {"$in": brand_ids}
        async for d in db.daily_sales.find(sales_match):
            outlet = outlets_map.get(d.get("outlet_id"))
            brand_id = (outlet or {}).get("brand_id") or d.get("brand_id")
            if brand_ids and brand_id not in brand_ids:
                continue
            month = (d.get("sales_date") or "")[:7]
            dim_key = tuple(_dim_value(dim, d.get("outlet_id"), brand_id, None, None, month) for dim in dimensions)
            row = agg.setdefault(dim_key, {})
            row["sales"] = row.get("sales", 0.0) + float(d.get("grand_total", 0) or 0)
            row["transaction_count"] = row.get("transaction_count", 0.0) + float(d.get("transaction_count", 0) or 0)

    # ---- Purchase value / PO count / GR count / AP exposure ----
    if any(m in metrics for m in ("purchase_value", "po_count", "gr_count", "ap_exposure")):
        gr_match: dict = {"deleted_at": None}
        if period_from or period_to:
            gr_match["receive_date"] = {}
            if period_from:
                gr_match["receive_date"]["$gte"] = period_from
            if period_to:
                gr_match["receive_date"]["$lte"] = period_to
        if vendor_ids:
            gr_match["vendor_id"] = {"$in": vendor_ids}
        if outlet_ids:
            gr_match["outlet_id"] = {"$in": outlet_ids}
        async for gr in db.goods_receipts.find(gr_match):
            vendor_id = gr.get("vendor_id")
            outlet_id = gr.get("outlet_id")
            outlet = outlets_map.get(outlet_id)
            brand_id = (outlet or {}).get("brand_id")
            if brand_ids and brand_id not in brand_ids:
                continue
            month = (gr.get("receive_date") or "")[:7]

            grand = float(gr.get("grand_total", 0) or 0)
            unpaid = not (gr.get("paid_at") or gr.get("payment_status") == "paid")

            # If category dimension requested → split by line.item.category
            if "category" in dimensions:
                lines = gr.get("lines", [])
                # Split grand_total weighted by line total
                line_totals = [float(ln.get("total", 0) or 0) for ln in lines]
                line_sum = sum(line_totals) or 1
                for ln, lt in zip(lines, line_totals):
                    cat_id = (items_map.get(ln.get("item_id"), {}) or {}).get("category_id")
                    if category_ids and cat_id not in category_ids:
                        continue
                    weight = lt / line_sum
                    dim_key = tuple(_dim_value(dim, outlet_id, brand_id, vendor_id, cat_id, month) for dim in dimensions)
                    r = agg.setdefault(dim_key, {})
                    r["purchase_value"] = r.get("purchase_value", 0.0) + grand * weight
                    r["po_count"] = r.get("po_count", 0.0)  # po_count is per PO - skip line split
                    r["gr_count"] = r.get("gr_count", 0.0) + (1 / max(len(lines), 1))
                    if unpaid:
                        r["ap_exposure"] = r.get("ap_exposure", 0.0) + grand * weight
            else:
                if category_ids:
                    continue  # filter doesn't apply at GR level
                dim_key = tuple(_dim_value(dim, outlet_id, brand_id, vendor_id, None, month) for dim in dimensions)
                r = agg.setdefault(dim_key, {})
                r["purchase_value"] = r.get("purchase_value", 0.0) + grand
                r["gr_count"] = r.get("gr_count", 0.0) + 1
                if unpaid:
                    r["ap_exposure"] = r.get("ap_exposure", 0.0) + grand

        # PO count
        if "po_count" in metrics:
            po_match: dict = {"deleted_at": None}
            if period_from or period_to:
                po_match["order_date"] = {}
                if period_from:
                    po_match["order_date"]["$gte"] = period_from
                if period_to:
                    po_match["order_date"]["$lte"] = period_to
            if vendor_ids:
                po_match["vendor_id"] = {"$in": vendor_ids}
            if outlet_ids:
                po_match["outlet_id"] = {"$in": outlet_ids}
            async for po in db.purchase_orders.find(po_match):
                vendor_id = po.get("vendor_id")
                outlet_id = po.get("outlet_id")
                outlet = outlets_map.get(outlet_id)
                brand_id = (outlet or {}).get("brand_id")
                month = (po.get("order_date") or "")[:7]
                if "category" in dimensions:
                    continue  # category dim not meaningful at PO header
                dim_key = tuple(_dim_value(dim, outlet_id, brand_id, vendor_id, None, month) for dim in dimensions)
                r = agg.setdefault(dim_key, {})
                r["po_count"] = r.get("po_count", 0.0) + 1

    # ---- COGS / Gross Profit (JE side) ----
    if any(m in metrics for m in ("cogs", "gross_profit")):
        # Find COGS COA ids
        cogs_coa_ids: list[str] = []
        async for c in db.chart_of_accounts.find({"type": "cogs", "is_postable": True, "deleted_at": None}):
            cogs_coa_ids.append(c["id"])
        if cogs_coa_ids:
            je_match: dict = {"deleted_at": None, "status": "posted"}
            if period_from or period_to:
                je_match["entry_date"] = {}
                if period_from:
                    je_match["entry_date"]["$gte"] = period_from
                if period_to:
                    je_match["entry_date"]["$lte"] = period_to
            async for je in db.journal_entries.find(je_match):
                month = (je.get("entry_date") or je.get("period") or "")[:7]
                for ln in je.get("lines", []):
                    if ln.get("coa_id") not in cogs_coa_ids:
                        continue
                    cogs_amt = float(ln.get("dr", 0) or 0) - float(ln.get("cr", 0) or 0)
                    if cogs_amt == 0:
                        continue
                    outlet_id = ln.get("dim_outlet")
                    if outlet_ids and outlet_id not in outlet_ids:
                        continue
                    outlet = outlets_map.get(outlet_id)
                    brand_id = (outlet or {}).get("brand_id") or ln.get("dim_brand")
                    if brand_ids and brand_id not in brand_ids:
                        continue
                    dim_key = tuple(_dim_value(dim, outlet_id, brand_id, None, None, month) for dim in dimensions)
                    r = agg.setdefault(dim_key, {})
                    r["cogs"] = r.get("cogs", 0.0) + cogs_amt

    # Compute derived gross_profit & finalize
    rows_out: list[dict[str, Any]] = []
    grand_totals: dict[str, float] = {m: 0.0 for m in metrics}
    for dim_key, vals in agg.items():
        if "gross_profit" in metrics:
            vals["gross_profit"] = vals.get("sales", 0) - vals.get("cogs", 0)
        # Extract only requested metrics
        out_row: dict[str, Any] = {}
        for i, dim in enumerate(dimensions):
            label_id = dim_key[i]
            label = label_id
            if dim == "outlet":
                label = outlets_map.get(label_id, {}).get("name", label_id) if label_id else "(tanpa outlet)"
            elif dim == "brand":
                label = brands_map.get(label_id, {}).get("name", label_id) if label_id else "(tanpa brand)"
            elif dim == "vendor":
                label = vendors_map.get(label_id, {}).get("name", label_id) if label_id else "(tanpa vendor)"
            elif dim == "category":
                label = categories_map.get(label_id, {}).get("name", label_id) if label_id else "(tanpa kategori)"
            elif dim == "month":
                label = label_id or "(no date)"
            out_row[f"dim_{dim}"] = label
            out_row[f"dim_{dim}_id"] = label_id
        for m in metrics:
            v = round(vals.get(m, 0.0), 2)
            out_row[m] = v
            grand_totals[m] = grand_totals.get(m, 0.0) + v
        rows_out.append(out_row)

    # Sort
    if sort_by:
        if sort_by in metrics or sort_by.startswith("dim_"):
            rows_out.sort(key=lambda r: (r.get(sort_by) or 0), reverse=(sort_dir.lower() == "desc"))
    else:
        # Default: sort by first metric desc
        first_metric = metrics[0]
        rows_out.sort(key=lambda r: r.get(first_metric, 0), reverse=True)

    if limit and len(rows_out) > limit:
        rows_out = rows_out[:limit]

    return {
        "rows": rows_out,
        "totals": {m: round(grand_totals.get(m, 0), 2) for m in metrics},
        "row_count": len(rows_out),
        "config": {
            "dimensions": dimensions,
            "metrics": metrics,
            "filters": {
                "period_from": period_from, "period_to": period_to,
                "outlet_ids": outlet_ids, "brand_ids": brand_ids,
                "vendor_ids": vendor_ids, "category_ids": category_ids,
            },
            "sort_by": sort_by, "sort_dir": sort_dir,
        },
    }


def _dim_value(dim: str, outlet_id, brand_id, vendor_id, category_id, month) -> Optional[str]:
    if dim == "outlet":
        return outlet_id
    if dim == "brand":
        return brand_id
    if dim == "vendor":
        return vendor_id
    if dim == "category":
        return category_id
    if dim == "month":
        return month
    return None


# ============================================================
# 3. PIVOT MATRIX
# ============================================================
async def pivot_matrix(
    *,
    dim_x: str,
    dim_y: str,
    metric: str,
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
) -> dict[str, Any]:
    """Pivot 2-D matrix using report_builder under the hood.
    Output: { x_labels: [...], y_labels: [...], cells: [[...]] }.
    """
    if dim_x == dim_y:
        raise ValidationError("dim_x dan dim_y harus berbeda")
    result = await report_builder(
        dimensions=[dim_y, dim_x],
        metrics=[metric],
        period_from=period_from, period_to=period_to,
        limit=10000,
    )
    rows = result["rows"]
    x_labels: list[str] = []
    y_labels: list[str] = []
    matrix: dict[tuple[str, str], float] = {}
    for r in rows:
        xl = r.get(f"dim_{dim_x}", "(–)")
        yl = r.get(f"dim_{dim_y}", "(–)")
        if xl not in x_labels:
            x_labels.append(xl)
        if yl not in y_labels:
            y_labels.append(yl)
        matrix[(yl, xl)] = float(r.get(metric, 0) or 0)
    x_labels.sort()
    y_labels.sort()
    cells: list[list[float]] = []
    row_totals: list[float] = []
    col_totals: dict[str, float] = {x: 0.0 for x in x_labels}
    grand: float = 0.0
    for yl in y_labels:
        row: list[float] = []
        rt = 0.0
        for xl in x_labels:
            v = round(matrix.get((yl, xl), 0.0), 2)
            row.append(v)
            rt += v
            col_totals[xl] = round(col_totals[xl] + v, 2)
        cells.append(row)
        row_totals.append(round(rt, 2))
        grand += rt
    return {
        "dim_x": dim_x, "dim_y": dim_y, "metric": metric,
        "x_labels": x_labels, "y_labels": y_labels,
        "cells": cells,
        "row_totals": row_totals,
        "col_totals": [col_totals[x] for x in x_labels],
        "grand_total": round(grand, 2),
        "filters": {"period_from": period_from, "period_to": period_to},
    }


# ============================================================
# 4. COMPARATIVES (MoM / YoY)
# ============================================================
async def comparatives(
    *,
    metric: str,
    period: str,  # YYYY-MM
    compare_to: str = "mom",  # mom | yoy
    outlet_ids: Optional[list[str]] = None,
    brand_ids: Optional[list[str]] = None,
) -> dict[str, Any]:
    """Compare metric value between current period and previous (mom) or last year (yoy).
    Also returns 12-month rolling trend.
    """
    if metric not in SUPPORTED_METRICS:
        raise ValidationError(f"Metrik tidak didukung: {metric}")
    if compare_to not in ("mom", "yoy"):
        raise ValidationError("compare_to harus 'mom' atau 'yoy'")
    try:
        y, m = [int(x) for x in period.split("-")]
        assert 1 <= m <= 12
    except Exception as e:
        raise ValidationError("period harus YYYY-MM") from e

    if compare_to == "mom":
        prev_y, prev_m = (y, m - 1) if m > 1 else (y - 1, 12)
    else:  # yoy
        prev_y, prev_m = y - 1, m
    prev_period = f"{prev_y:04d}-{prev_m:02d}"

    cur = await _metric_for_period(metric, period, outlet_ids, brand_ids)
    prev = await _metric_for_period(metric, prev_period, outlet_ids, brand_ids)

    # 12-month rolling (current and 11 before, including prev as needed)
    rolling: list[dict[str, Any]] = []
    for i in range(11, -1, -1):
        cur_y = y
        cur_m = m - i
        while cur_m <= 0:
            cur_m += 12
            cur_y -= 1
        per = f"{cur_y:04d}-{cur_m:02d}"
        v = await _metric_for_period(metric, per, outlet_ids, brand_ids)
        rolling.append({"period": per, "value": round(v, 2)})

    delta = cur - prev
    delta_pct = round((delta / prev) * 100, 2) if prev else None

    return {
        "metric": metric,
        "period": period,
        "compare_to": compare_to,
        "current": round(cur, 2),
        "previous_period": prev_period,
        "previous": round(prev, 2),
        "delta": round(delta, 2),
        "delta_pct": delta_pct,
        "rolling_12m": rolling,
        "filters": {"outlet_ids": outlet_ids, "brand_ids": brand_ids},
    }


async def _metric_for_period(metric: str, period: str,
                              outlet_ids: Optional[list[str]],
                              brand_ids: Optional[list[str]]) -> float:
    """Return a single scalar for a given metric in a given period (YYYY-MM)."""
    period_from = f"{period}-01"
    y, m = [int(x) for x in period.split("-")]
    if m == 12:
        ny, nm = y + 1, 1
    else:
        ny, nm = y, m + 1
    period_to = f"{ny:04d}-{nm:02d}-01"

    # Use report_builder w/o dim → all aggregated
    result = await report_builder(
        dimensions=["month"],  # ensure month-aggregated
        metrics=[metric],
        period_from=period_from, period_to=period_to,
        outlet_ids=outlet_ids, brand_ids=brand_ids,
        limit=100,
    )
    return float(result["totals"].get(metric, 0) or 0)


# ============================================================
# 5. SAVED REPORT DEFINITIONS (per user)
# ============================================================
async def save_report(
    *, user_id: str, name: str, description: Optional[str] = None,
    config: dict[str, Any], saved_id: Optional[str] = None,
    public: bool = False,
) -> dict[str, Any]:
    if not name or not name.strip():
        raise ValidationError("Nama report wajib")
    if not config or not isinstance(config, dict):
        raise ValidationError("Config report wajib")
    db = get_db()
    if saved_id:
        existing = await db.saved_reports.find_one({"id": saved_id, "deleted_at": None})
        if not existing:
            raise NotFoundError("Saved report tidak ditemukan")
        if existing["owner_user_id"] != user_id:
            raise ValidationError("Hanya owner yang boleh edit report ini")
        await db.saved_reports.update_one(
            {"id": saved_id},
            {"$set": {
                "name": name.strip(),
                "description": (description or "").strip() or None,
                "config": config,
                "public": bool(public),
                "updated_at": _now(),
                "updated_by": user_id,
            }},
        )
        return await get_saved(saved_id, user_id=user_id)

    doc = {
        "id": str(uuid.uuid4()),
        "owner_user_id": user_id,
        "name": name.strip(),
        "description": (description or "").strip() or None,
        "config": config,
        "public": bool(public),
        "created_at": _now(), "updated_at": _now(),
        "created_by": user_id, "updated_by": user_id,
        "deleted_at": None,
    }
    await db.saved_reports.insert_one(doc)
    return serialize(doc)


async def list_saved_reports(*, user_id: str) -> list[dict]:
    db = get_db()
    q = {"deleted_at": None, "$or": [{"owner_user_id": user_id}, {"public": True}]}
    items = await db.saved_reports.find(q).sort([("updated_at", -1)]).to_list(500)
    return [serialize(d) for d in items]


async def get_saved(saved_id: str, *, user_id: str) -> dict:
    db = get_db()
    d = await db.saved_reports.find_one({"id": saved_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Saved report tidak ditemukan")
    if d.get("owner_user_id") != user_id and not d.get("public"):
        raise ValidationError("Tidak ada akses ke report ini")
    return serialize(d)


async def delete_saved(saved_id: str, *, user_id: str) -> dict:
    db = get_db()
    d = await db.saved_reports.find_one({"id": saved_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Saved report tidak ditemukan")
    if d.get("owner_user_id") != user_id:
        raise ValidationError("Hanya owner yang boleh hapus")
    await db.saved_reports.update_one(
        {"id": saved_id},
        {"$set": {"deleted_at": _now(), "updated_at": _now(), "updated_by": user_id}},
    )
    return {"id": saved_id, "deleted": True}


# ============================================================
# 6. CATALOG (helper for UI)
# ============================================================
def get_catalog() -> dict[str, Any]:
    return {
        "dimensions": [
            {"key": k, "label": k.replace("_", " ").title()} for k in SUPPORTED_DIMENSIONS
        ],
        "metrics": [
            {"key": k, "label": v} for k, v in SUPPORTED_METRICS.items()
        ],
        "comparatives": [
            {"key": "mom", "label": "Month-over-Month"},
            {"key": "yoy", "label": "Year-over-Year"},
        ],
    }



# ============================================================
# 7. EXCEL EXPORTS (Phase 4.1 — Sales & Outlet Reports)
# ============================================================

async def generate_daily_sales_excel(
    *,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    outlet_ids: Optional[list[str]] = None,
    brand_ids: Optional[list[str]] = None,
):
    """Generate Daily Sales Summary Excel report.
    
    Columns: Date, Outlet, Brand, Grand Total, Transaction Count, Status
    Features: Subtotals per outlet, Grand total row, Styling
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        write_total_row,
        autosize_columns,
        freeze_header_row,
        apply_currency_format,
        format_date,
    )
    
    db = get_db()
    
    # Build query
    match_filter: dict = {"deleted_at": None, "status": "validated"}
    if date_from or date_to:
        match_filter["sales_date"] = {}
        if date_from:
            match_filter["sales_date"]["$gte"] = date_from
        if date_to:
            match_filter["sales_date"]["$lte"] = date_to
    if outlet_ids:
        match_filter["outlet_id"] = {"$in": outlet_ids}
    if brand_ids:
        match_filter["brand_id"] = {"$in": brand_ids}
    
    # Load data
    sales_docs = []
    async for doc in db.daily_sales.find(match_filter).sort([("sales_date", 1), ("outlet_id", 1)]):
        sales_docs.append(doc)
    
    # Load lookups
    outlet_map = {}
    async for o in db.outlets.find({"deleted_at": None}):
        outlet_map[o["id"]] = o.get("name", o["id"])
    
    brand_map = {}
    async for b in db.brands.find({"deleted_at": None}):
        brand_map[b["id"]] = b.get("name", b["id"])
    
    # Create workbook
    wb = create_workbook("Daily Sales")
    ws = wb.active
    
    # Add header
    subtitle = f"Period: {format_date(date_from) if date_from else 'All'} - {format_date(date_to) if date_to else 'All'}"
    current_row = add_report_header(
        ws,
        title="Daily Sales Summary Report",
        subtitle=subtitle,
    )
    
    # Table headers
    headers = ["Date", "Outlet", "Brand", "Grand Total", "Transaction Count", "Status"]
    current_row = write_table_headers(ws, headers, current_row)
    
    # Data rows
    grand_total = 0.0
    grand_trx_count = 0
    
    for doc in sales_docs:
        outlet_name = outlet_map.get(doc.get("outlet_id"), doc.get("outlet_id") or "")
        brand_name = brand_map.get(doc.get("brand_id"), doc.get("brand_id") or "")
        sales_total = float(doc.get("grand_total", 0) or 0)
        trx_count = int(doc.get("transaction_count", 0) or 0)
        
        row_data = [
            format_date(doc.get("sales_date")),
            outlet_name,
            brand_name,
            sales_total,
            trx_count,
            doc.get("status", ""),
        ]
        write_data_row(ws, row_data, current_row, number_cols=[4, 5])
        
        grand_total += sales_total
        grand_trx_count += trx_count
        current_row += 1
    
    # Total row
    if sales_docs:
        total_row = ["", "", "TOTAL", grand_total, grand_trx_count, ""]
        write_total_row(ws, total_row, current_row, number_cols=[4, 5])
    
    # Styling
    apply_currency_format(ws, 4, current_row - len(sales_docs), current_row, currency="Rp")
    autosize_columns(ws)
    freeze_header_row(ws, header_row=current_row - len(sales_docs) - 1)
    
    return wb


async def generate_outlet_performance_excel(
    *,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    outlet_ids: Optional[list[str]] = None,
):
    """Generate Outlet Performance Excel report.
    
    Columns: Outlet, Total Sales, Avg Daily Sales, Transaction Count, Days Active
    Features: Bar chart comparison, Conditional formatting
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        write_total_row,
        autosize_columns,
        freeze_header_row,
        apply_currency_format,
        format_date,
        add_bar_chart,
    )
    
    db = get_db()
    
    # Build query
    match_filter: dict = {"deleted_at": None, "status": "validated"}
    if date_from or date_to:
        match_filter["sales_date"] = {}
        if date_from:
            match_filter["sales_date"]["$gte"] = date_from
        if date_to:
            match_filter["sales_date"]["$lte"] = date_to
    if outlet_ids:
        match_filter["outlet_id"] = {"$in": outlet_ids}
    
    # Aggregate per outlet
    pipeline = [
        {"$match": match_filter},
        {
            "$group": {
                "_id": "$outlet_id",
                "total_sales": {"$sum": "$grand_total"},
                "trx_count": {"$sum": "$transaction_count"},
                "days_count": {"$sum": 1},
            }
        },
        {"$sort": {"total_sales": -1}},
    ]
    
    results = []
    async for doc in db.daily_sales.aggregate(pipeline):
        results.append(doc)
    
    # Load outlet names
    outlet_map = {}
    async for o in db.outlets.find({"deleted_at": None}):
        outlet_map[o["id"]] = o.get("name", o["id"])
    
    # Create workbook
    wb = create_workbook("Outlet Performance")
    ws = wb.active
    
    # Add header
    subtitle = f"Period: {format_date(date_from) if date_from else 'All'} - {format_date(date_to) if date_to else 'All'}"
    current_row = add_report_header(
        ws,
        title="Outlet Performance Report",
        subtitle=subtitle,
    )
    
    # Table headers
    headers = ["Outlet", "Total Sales", "Days Active", "Avg Daily Sales", "Transaction Count"]
    current_row = write_table_headers(ws, headers, current_row)
    data_start_row = current_row
    
    # Data rows
    grand_total = 0.0
    grand_trx = 0
    
    for doc in results:
        outlet_name = outlet_map.get(doc["_id"], doc["_id"] or "Unknown")
        total_sales = float(doc.get("total_sales", 0) or 0)
        days = int(doc.get("days_count", 0) or 0)
        avg_daily = total_sales / days if days > 0 else 0
        trx_count = int(doc.get("trx_count", 0) or 0)
        
        row_data = [
            outlet_name,
            total_sales,
            days,
            avg_daily,
            trx_count,
        ]
        write_data_row(ws, row_data, current_row, number_cols=[2, 3, 4, 5])
        
        grand_total += total_sales
        grand_trx += trx_count
        current_row += 1
    
    # Total row
    if results:
        total_days = sum(int(r.get("days_count", 0) or 0) for r in results)
        total_row = ["TOTAL", grand_total, total_days, grand_total / total_days if total_days > 0 else 0, grand_trx]
        write_total_row(ws, total_row, current_row, number_cols=[2, 3, 4, 5])
    
    # Styling
    apply_currency_format(ws, 2, data_start_row, current_row, currency="Rp")
    apply_currency_format(ws, 4, data_start_row, current_row, currency="Rp")
    autosize_columns(ws)
    freeze_header_row(ws, header_row=data_start_row - 1)
    
    # Add bar chart (Total Sales)
    if results and len(results) > 1:
        data_range = f"B{data_start_row}:B{current_row - 1}"
        cats_range = f"A{data_start_row}:A{current_row - 1}"
        add_bar_chart(ws, "Total Sales by Outlet", data_range, cats_range, position="G2")
    
    return wb


async def generate_fdo_history_excel(
    *,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    outlet_ids: Optional[list[str]] = None,
    status: Optional[str] = None,
):
    """Generate FDO History Excel report.
    
    Columns: Doc No, Date, Outlet, Items Summary, Status, Approved By
    Features: Status color coding
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        autosize_columns,
        freeze_header_row,
        format_date,
        format_datetime,
    )
    from openpyxl.styles import PatternFill
    
    db = get_db()
    
    # Build query
    match_filter: dict = {"deleted_at": None, "request_type": "fdo"}
    if date_from or date_to:
        match_filter["request_date"] = {}
        if date_from:
            match_filter["request_date"]["$gte"] = date_from
        if date_to:
            match_filter["request_date"]["$lte"] = date_to
    if outlet_ids:
        match_filter["outlet_id"] = {"$in": outlet_ids}
    if status:
        match_filter["status"] = status
    
    # Load data
    fdo_docs = []
    async for doc in db.kdo_bdo_requests.find(match_filter).sort([("request_date", -1)]):
        fdo_docs.append(doc)
    
    # Load lookups
    outlet_map = {}
    async for o in db.outlets.find({"deleted_at": None}):
        outlet_map[o["id"]] = o.get("name", o["id"])
    
    user_map = {}
    async for u in db.users.find({"deleted_at": None}):
        user_map[u["id"]] = u.get("name", u.get("email", u["id"]))
    
    # Create workbook
    wb = create_workbook("FDO History")
    ws = wb.active
    
    # Add header
    subtitle = f"Period: {format_date(date_from) if date_from else 'All'} - {format_date(date_to) if date_to else 'All'}"
    current_row = add_report_header(
        ws,
        title="FDO History Report",
        subtitle=subtitle,
    )
    
    # Table headers
    headers = ["Doc No", "Request Date", "Outlet", "Items Count", "Status", "Approved By", "Approved At"]
    current_row = write_table_headers(ws, headers, current_row)
    data_start_row = current_row
    
    # Data rows
    for doc in fdo_docs:
        outlet_name = outlet_map.get(doc.get("outlet_id"), doc.get("outlet_id") or "")
        items_count = len(doc.get("lines", []))
        status_val = doc.get("status", "draft")
        approved_by_name = user_map.get(doc.get("approved_by"), "") if doc.get("approved_by") else ""
        approved_at = format_datetime(doc.get("approved_at")) if doc.get("approved_at") else ""
        
        row_data = [
            doc.get("doc_no", ""),
            format_date(doc.get("request_date")),
            outlet_name,
            items_count,
            status_val,
            approved_by_name,
            approved_at,
        ]
        write_data_row(ws, row_data, current_row, number_cols=[4])
        
        # Color code status cell
        status_cell = ws.cell(row=current_row, column=5)
        if status_val == "approved":
            status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        elif status_val == "rejected":
            status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        elif status_val == "pending":
            status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        current_row += 1
    
    # Styling
    autosize_columns(ws)
    freeze_header_row(ws, header_row=data_start_row - 1)
    
    return wb
