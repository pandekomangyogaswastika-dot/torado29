"""HR Service — Phase 5 + Sprint G.

Modules covered:
- Employee Advances (kasbon)
- Service Charge per period/outlet
- Incentive schemes + runs
- Vouchers (issue + redeem)
- FOC (free of charge)
- LB Fund ledger
- Payroll cycle (Sprint G enhanced: BPJS + PPh21)
- Salary Master (Sprint G)
- HR dashboard

Conventions:
- All collection ids are UUID4 strings.
- All datetimes in UTC ISO format.
- RBAC-checked at router layer; this service only enforces business invariants.
- Side-effects (journal posting) routed through services.journal_service.
"""
import csv
import io
import uuid
from datetime import datetime, timezone, timedelta
from typing import Optional

from core.audit import log as audit_log
from core.db import get_db, serialize
from core.exceptions import (
    ConflictError, NotFoundError, ValidationError,
)
from services import approval_service, journal_service
from utils.number_series import next_doc_no

# ── Sprint G: BPJS & PPh21 constants ──────────────────────────────────────────
BPJS_RATES = {
    "jkk_employer": 0.0054,   # Kecelakaan Kerja — employer only
    "jkm_employer": 0.0030,   # Kematian — employer only
    "jht_employer": 0.0370,   # Hari Tua employer share
    "jht_employee": 0.0200,   # Hari Tua employee share
    "jp_employer":  0.0200,   # Pensiun employer
    "jp_employee":  0.0100,   # Pensiun employee
    "jkes_employer": 0.0400,  # Kesehatan employer
    "jkes_employee": 0.0100,  # Kesehatan employee
    "jp_base_cap":   9_559_600,   # Max base salary for JP 2025
    "jkes_base_cap": 12_000_000,  # Max base salary for JKes
}

# PPh 21 brackets (UU HPP No.7/2021)
PPH21_BRACKETS = [
    (0,            60_000_000,   0.05),
    (60_000_000,   250_000_000,  0.15),
    (250_000_000,  500_000_000,  0.25),
    (500_000_000,  5_000_000_000, 0.30),
    (5_000_000_000, float("inf"), 0.35),
]

# PTKP 2026
PTKP_BY_STATUS = {
    "TK/0": 54_000_000, "TK/1": 58_500_000, "TK/2": 63_000_000, "TK/3": 67_500_000,
    "K/0":  58_500_000, "K/1":  63_000_000, "K/2":  67_500_000, "K/3":  72_000_000,
    "K/I/0": 112_500_000, "K/I/1": 117_000_000, "K/I/2": 121_500_000, "K/I/3": 126_000_000,
}
PTKP_DEFAULT = 54_000_000  # TK/0


def _compute_progressive_tax(pkp: float) -> float:
    """Compute annual PPh 21 using progressive brackets."""
    tax = 0.0
    remaining = pkp
    for lower, upper, rate in PPH21_BRACKETS:
        if remaining <= 0:
            break
        bracket_size = (upper - lower) if upper != float("inf") else remaining
        taxable_in_bracket = min(remaining, bracket_size)
        if taxable_in_bracket > 0:
            tax += taxable_in_bracket * rate
        remaining -= taxable_in_bracket
    return round(tax, 2)


def _calc_bpjs(gross: float, enrolled: bool) -> dict:
    """Return BPJS employee + employer breakdown for a given gross salary."""
    if not enrolled or gross <= 0:
        return {"employee": 0.0, "employer": 0.0, "detail": {}}
    r = BPJS_RATES
    jp_base  = min(gross, r["jp_base_cap"])
    jkes_base = min(gross, r["jkes_base_cap"])
    emp = round(
        gross    * r["jht_employee"] +
        jp_base  * r["jp_employee"]  +
        jkes_base* r["jkes_employee"],
        2,
    )
    er = round(
        gross    * r["jkk_employer"] +
        gross    * r["jkm_employer"] +
        gross    * r["jht_employer"] +
        jp_base  * r["jp_employer"]  +
        jkes_base* r["jkes_employer"],
        2,
    )
    return {
        "employee": emp, "employer": er,
        "detail": {
            "jht_employee": round(gross * r["jht_employee"], 2),
            "jp_employee":  round(jp_base * r["jp_employee"], 2),
            "jkes_employee": round(jkes_base * r["jkes_employee"], 2),
            "jkk_employer": round(gross * r["jkk_employer"], 2),
            "jkm_employer": round(gross * r["jkm_employer"], 2),
            "jht_employer": round(gross * r["jht_employer"], 2),
            "jp_employer":  round(jp_base * r["jp_employer"], 2),
            "jkes_employer": round(jkes_base * r["jkes_employer"], 2),
        },
    }


def _calc_pph21(monthly_gross: float, ptkp_status: str) -> dict:
    """Return monthly PPh 21 for an employee (gross method)."""
    if monthly_gross <= 0:
        return {"monthly_tax": 0.0, "annual_gross": 0.0, "annual_pkp": 0.0, "annual_tax": 0.0}
    annual_gross = monthly_gross * 12
    biaya_jabatan = min(annual_gross * 0.05, 6_000_000)
    net_income = annual_gross - biaya_jabatan
    ptkp = PTKP_BY_STATUS.get(ptkp_status, PTKP_DEFAULT)
    annual_pkp = max(net_income - ptkp, 0)
    annual_pkp = int(annual_pkp / 1000) * 1000  # round down to nearest 1000
    annual_tax = _compute_progressive_tax(annual_pkp)
    monthly_tax = round(annual_tax / 12, 2)
    return {
        "monthly_tax": monthly_tax,
        "annual_gross": annual_gross,
        "biaya_jabatan": biaya_jabatan,
        "ptkp": ptkp,
        "annual_pkp": annual_pkp,
        "annual_tax": annual_tax,
        "ptkp_status": ptkp_status,
    }


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _today() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def _period_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m")


# ============================================================
# EMPLOYEE ADVANCES (Kasbon)
# ============================================================

def _build_schedule(principal: float, terms_months: int, start_date: str) -> tuple[float, list[dict]]:
    """Build flat amortization schedule (equal installments)."""
    if terms_months <= 0:
        terms_months = 1
    monthly = round(principal / terms_months, 2)
    # Last installment absorbs rounding diff
    schedule: list[dict] = []
    remaining = principal
    try:
        anchor = datetime.strptime(start_date, "%Y-%m-%d")
    except Exception:
        anchor = datetime.now(timezone.utc)
    for i in range(terms_months):
        # Due date: same day next i+1 months (clip to month end if needed)
        y = anchor.year
        m = anchor.month + i + 1
        while m > 12:
            m -= 12
            y += 1
        d = anchor.day
        # Clip d to last day of month
        if m == 2 and d > 28:
            d = 28
        elif m in (4, 6, 9, 11) and d > 30:
            d = 30
        try:
            due = datetime(y, m, d).strftime("%Y-%m-%d")
        except Exception:
            due = f"{y:04d}-{m:02d}-28"
        amount = monthly if i < terms_months - 1 else round(remaining, 2)
        remaining -= amount
        schedule.append({
            "period": f"{y:04d}-{m:02d}",
            "due_date": due,
            "amount": amount,
            "paid": False,
            "paid_at": None,
        })
    return monthly, schedule


async def list_advances(
    *, employee_id: Optional[str] = None, outlet_id: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1, per_page: int = 20,
):
    db = get_db()
    q: dict = {"deleted_at": None}
    if employee_id:
        q["employee_id"] = employee_id
    if outlet_id:
        q["outlet_id"] = outlet_id
    if status:
        q["status"] = status
    skip = (page - 1) * per_page
    items = await db.employee_advances.find(q).sort([("advance_date", -1), ("created_at", -1)]).skip(skip).limit(per_page).to_list(per_page)
    total = await db.employee_advances.count_documents(q)
    # Enrich with employee name
    emp_ids = list({d.get("employee_id") for d in items if d.get("employee_id")})
    emp_map: dict = {}
    if emp_ids:
        async for e in db.employees.find({"id": {"$in": emp_ids}}):
            emp_map[e["id"]] = e.get("full_name", e["id"])
    out = []
    for d in items:
        s = serialize(d)
        s["employee_name"] = emp_map.get(s.get("employee_id"))
        out.append(s)
    return out, {"page": page, "per_page": per_page, "total": total}


async def get_advance(adv_id: str) -> dict:
    db = get_db()
    d = await db.employee_advances.find_one({"id": adv_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Employee advance tidak ditemukan")
    s = serialize(d)
    if s.get("employee_id"):
        emp = await db.employees.find_one({"id": s["employee_id"]})
        if emp:
            s["employee_name"] = emp.get("full_name")
    return s


async def create_advance(payload: dict, *, user: dict) -> dict:
    db = get_db()
    employee_id = payload.get("employee_id")
    if not employee_id:
        raise ValidationError("employee_id wajib")
    emp = await db.employees.find_one({"id": employee_id, "deleted_at": None})
    if not emp:
        raise ValidationError("Employee tidak ditemukan")

    principal = float(payload.get("principal", 0) or 0)
    if principal <= 0:
        raise ValidationError("principal harus > 0")
    terms_months = int(payload.get("terms_months", 1) or 1)
    if terms_months < 1 or terms_months > 24:
        raise ValidationError("terms_months 1..24")
    advance_date = payload.get("advance_date") or _today()
    monthly, schedule = _build_schedule(principal, terms_months, advance_date)
    doc_no = await next_doc_no("EA")
    doc = {
        "id": str(uuid.uuid4()),
        "doc_no": doc_no,
        "employee_id": employee_id,
        "outlet_id": payload.get("outlet_id") or emp.get("outlet_id"),
        "advance_date": advance_date,
        "principal": principal,
        "terms_months": terms_months,
        "monthly_installment": monthly,
        "schedule": schedule,
        "status": "draft",
        "reason": payload.get("reason"),
        "payment_method_id": payload.get("payment_method_id"),
        "approved_by": None, "approved_at": None,
        "disbursed_at": None,
        "journal_entry_id": None,
        "settled_at": None,
        "notes": payload.get("notes"),
        "created_at": _now(), "updated_at": _now(), "deleted_at": None,
        "created_by": user["id"],
    }
    await db.employee_advances.insert_one(doc)
    await audit_log(user_id=user["id"], entity_type="employee_advance",
                    entity_id=doc["id"], action="create")
    return await get_advance(doc["id"])


async def submit_advance_for_approval(adv_id: str, *, user: dict) -> dict:
    """Move EA from draft → awaiting_approval (engine flow starts)."""
    db = get_db()
    d = await db.employee_advances.find_one({"id": adv_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Employee advance tidak ditemukan")
    if d["status"] != "draft":
        raise ValidationError(f"Hanya draft yang bisa di-submit. Status saat ini: {d['status']}")
    await db.employee_advances.update_one(
        {"id": adv_id},
        {"$set": {"status": "awaiting_approval", "submitted_at": _now(), "updated_at": _now()}},
    )
    await audit_log(user_id=user["id"], entity_type="employee_advance",
                    entity_id=adv_id, action="submit")
    fresh = await db.employee_advances.find_one({"id": adv_id})
    fresh_s = serialize(fresh)
    try:
        state = await approval_service.evaluate("employee_advance", fresh_s)
        await approval_service.notify_pending_approvers(
            "employee_advance", fresh_s, state=state, triggered_by=user,
        )
    except Exception:  # noqa: BLE001
        pass
    return await get_advance(adv_id)


async def approve_advance(adv_id: str, *, user: dict, note: str | None = None) -> dict:
    """Multi-tier approve via engine. On final approval, disburse + post JE."""
    db = get_db()
    d = await db.employee_advances.find_one({"id": adv_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Employee advance tidak ditemukan")
    # Engine accepts draft / submitted / awaiting_approval
    if d["status"] not in ("draft", "submitted", "awaiting_approval"):
        raise ValidationError(f"Status saat ini: {d['status']} tidak bisa di-approve")

    # Detect workflow presence: when no workflow is configured, preserve LEGACY behavior
    state = await approval_service.evaluate("employee_advance", serialize(d))
    has_wf = bool(state.get("has_workflow"))

    if not has_wf:
        # LEGACY: single-step + immediate disburse
        if d["status"] != "draft":
            raise ValidationError(f"Status saat ini: {d['status']}, hanya draft yang bisa di-approve (legacy)")
        je = await journal_service.post_for_employee_advance(
            {**d, "disbursed_at": _now()}, user_id=user["id"],
        )
        await db.employee_advances.update_one(
            {"id": adv_id},
            {"$set": {
                "status": "repaying",
                "approved_at": _now(), "approved_by": user["id"],
                "disbursed_at": _now(),
                "journal_entry_id": je["id"] if je else None,
                "updated_at": _now(),
            }},
        )
        await audit_log(user_id=user["id"], entity_type="employee_advance",
                        entity_id=adv_id, action="approve")
        return await get_advance(adv_id)

    # WORKFLOW: delegate to engine; if completed → disburse
    res = await approval_service.approve("employee_advance", adv_id, user=user, note=note)
    entity = res["entity"]
    if entity.get("status") == "approved" and not entity.get("journal_entry_id"):
        je = await journal_service.post_for_employee_advance(
            {**entity, "disbursed_at": _now()}, user_id=user["id"],
        )
        await db.employee_advances.update_one(
            {"id": adv_id},
            {"$set": {
                "status": "repaying",
                "approved_at": _now(), "approved_by": user["id"],
                "disbursed_at": _now(),
                "journal_entry_id": je["id"] if je else None,
                "updated_at": _now(),
            }},
        )
    await audit_log(user_id=user["id"], entity_type="employee_advance",
                    entity_id=adv_id, action="approve_step")
    return await get_advance(adv_id)


async def reject_advance(adv_id: str, *, user: dict, reason: str) -> dict:
    res = await approval_service.reject("employee_advance", adv_id, user=user, reason=reason)
    return res["entity"]


async def get_advance_approval_state(adv_id: str) -> dict:
    db = get_db()
    d = await db.employee_advances.find_one({"id": adv_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Employee advance tidak ditemukan")
    return await approval_service.evaluate("employee_advance", serialize(d))


async def mark_advance_installment_paid(adv_id: str, period: str, *, user: dict) -> dict:
    """Mark a schedule line as paid (finance staff/HR action). Does not generate JE — payroll posting handles offset."""
    db = get_db()
    d = await db.employee_advances.find_one({"id": adv_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Employee advance tidak ditemukan")
    schedule = d.get("schedule", [])
    found = False
    for line in schedule:
        if line.get("period") == period and not line.get("paid"):
            line["paid"] = True
            line["paid_at"] = _now()
            found = True
            break
    if not found:
        raise ValidationError(f"Schedule line {period} tidak ditemukan / sudah paid")
    all_paid = all(item.get("paid") for item in schedule)
    update: dict = {"schedule": schedule, "updated_at": _now()}
    if all_paid:
        update["status"] = "settled"
        update["settled_at"] = _now()
    await db.employee_advances.update_one({"id": adv_id}, {"$set": update})
    await audit_log(user_id=user["id"], entity_type="employee_advance",
                    entity_id=adv_id, action="mark_paid",
                    after={"period": period})
    return await get_advance(adv_id)


# ============================================================
# SERVICE CHARGE
# ============================================================

async def list_service_charge(
    *, period: Optional[str] = None, outlet_id: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1, per_page: int = 20,
):
    db = get_db()
    q: dict = {"deleted_at": None}
    if period:
        q["period"] = period
    if outlet_id:
        q["outlet_id"] = outlet_id
    if status:
        q["status"] = status
    skip = (page - 1) * per_page
    items = await db.service_charge_periods.find(q).sort([("period", -1)]).skip(skip).limit(per_page).to_list(per_page)
    total = await db.service_charge_periods.count_documents(q)
    return [serialize(d) for d in items], {"page": page, "per_page": per_page, "total": total}


async def get_service_charge(sc_id: str) -> dict:
    db = get_db()
    d = await db.service_charge_periods.find_one({"id": sc_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Service charge period tidak ditemukan")
    s = serialize(d)
    # Enrich outlet name
    outlet = await db.outlets.find_one({"id": s.get("outlet_id")})
    s["outlet_name"] = outlet.get("name") if outlet else None
    return s


async def calculate_service_charge(payload: dict, *, user: dict) -> dict:
    """Calculate service charge for period/outlet:
    - Sum service_charge from validated daily_sales for outlet+period
    - Deduct LB% and LD% (defaults from `service_charge_policy` business rule
      resolved via outlet → brand → group hierarchy if not explicitly provided)
    - Distribute remainder by employee days_worked (default working-days from
      policy if not overridden in payload)
    """
    from services import business_rules_service  # local import to avoid cycle

    db = get_db()
    period = payload.get("period") or _period_now()
    outlet_id = payload.get("outlet_id")
    if not outlet_id:
        raise ValidationError("outlet_id wajib")
    outlet = await db.outlets.find_one({"id": outlet_id, "deleted_at": None})
    if not outlet:
        raise ValidationError("Outlet tidak ditemukan")

    # Resolve service_charge_policy if user did not override
    policy = await business_rules_service.resolve_rule(
        rule_type="service_charge_policy",
        outlet_id=outlet_id,
        brand_id=outlet.get("brand_id"),
        on_date=f"{period}-01",
    )
    policy_data = (policy or {}).get("rule_data") or {}

    lb_pct = float(
        payload.get("lb_pct") if payload.get("lb_pct") is not None else policy_data.get("lb_pct", 0.05)
    )
    ld_pct = float(
        payload.get("ld_pct") if payload.get("ld_pct") is not None else policy_data.get("ld_pct", 0)
    )
    default_days = int(payload.get("default_working_days") or policy_data.get("default_working_days") or 22)

    if lb_pct < 0 or lb_pct > 0.5:
        raise ValidationError("lb_pct di luar batas (0..0.5)")
    if ld_pct < 0 or ld_pct > 0.5:
        raise ValidationError("ld_pct di luar batas (0..0.5)")

    # Aggregate validated daily_sales service_charge for outlet+period
    agg = db.daily_sales.aggregate([
        {"$match": {
            "deleted_at": None, "status": "validated",
            "outlet_id": outlet_id,
            "sales_date": {"$gte": f"{period}-01", "$lte": f"{period}-31"},
        }},
        {"$group": {"_id": None,
                    "total": {"$sum": {"$ifNull": ["$service_charge", 0]}}}},
    ])
    res = await agg.to_list(1)
    gross_service = float(res[0]["total"]) if res else 0.0
    lb_amount = round(gross_service * lb_pct, 2)
    ld_amount = round(gross_service * ld_pct, 2)
    distributable = round(gross_service - lb_amount - ld_amount, 2)

    # Employees in outlet: assume 22 working days each by default; allow override via payload.allocations[]
    employees: list[dict] = []
    async for e in db.employees.find({"deleted_at": None, "outlet_id": outlet_id, "status": "active"}):
        employees.append(e)

    # Days worked override
    days_overrides = {a["employee_id"]: float(a.get("days_worked", default_days))
                      for a in payload.get("allocations", []) if a.get("employee_id")}
    total_days = sum(days_overrides.get(e["id"], default_days) for e in employees)
    allocations: list[dict] = []
    if total_days <= 0 or distributable <= 0 or not employees:
        for e in employees:
            allocations.append({
                "employee_id": e["id"],
                "employee_name": e.get("full_name"),
                "days_worked": days_overrides.get(e["id"], default_days),
                "share_pct": 0,
                "amount": 0,
            })
    else:
        for e in employees:
            d_w = days_overrides.get(e["id"], default_days)
            share = d_w / total_days
            amount = round(distributable * share, 2)
            allocations.append({
                "employee_id": e["id"],
                "employee_name": e.get("full_name"),
                "days_worked": d_w,
                "share_pct": round(share * 100, 2),
                "amount": amount,
            })

    # Upsert (one per period+outlet) — keep status flow safe
    existing = await db.service_charge_periods.find_one({
        "period": period, "outlet_id": outlet_id, "deleted_at": None,
    })
    common = {
        "period": period, "outlet_id": outlet_id,
        "brand_id": outlet.get("brand_id"),
        "gross_service": round(gross_service, 2),
        "lb_pct": lb_pct, "ld_pct": ld_pct,
        "lb_amount": lb_amount, "ld_amount": ld_amount,
        "distributable": distributable,
        "allocations": allocations,
        "status": "calculated",
        "calculated_at": _now(), "calculated_by": user["id"],
        "updated_at": _now(),
        "notes": payload.get("notes"),
        "policy_id": (policy or {}).get("id"),
        "policy_version": (policy or {}).get("version"),
        "policy_scope": ((policy or {}).get("scope_type"), (policy or {}).get("scope_id"))
            if policy else None,
    }
    if existing:
        if existing["status"] == "posted":
            raise ConflictError("Period sudah posted; tidak bisa di-recalculate")
        await db.service_charge_periods.update_one({"id": existing["id"]}, {"$set": common})
        sc_id = existing["id"]
    else:
        doc = {
            "id": str(uuid.uuid4()),
            "doc_no": f"SC-{period}-{outlet.get('code','')}",
            **common,
            "approved_at": None, "approved_by": None,
            "posted_at": None, "posted_by": None,
            "journal_entry_id": None,
            "created_at": _now(), "deleted_at": None,
            "created_by": user["id"],
        }
        await db.service_charge_periods.insert_one(doc)
        sc_id = doc["id"]
    await audit_log(user_id=user["id"], entity_type="service_charge",
                    entity_id=sc_id, action="calculate",
                    after={"period": period, "outlet_id": outlet_id})
    return await get_service_charge(sc_id)


async def approve_service_charge(sc_id: str, *, user: dict) -> dict:
    db = get_db()
    d = await db.service_charge_periods.find_one({"id": sc_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Service charge tidak ditemukan")
    if d["status"] != "calculated":
        raise ValidationError(f"Status saat ini: {d['status']}, hanya calculated yang bisa di-approve")
    await db.service_charge_periods.update_one(
        {"id": sc_id},
        {"$set": {"status": "approved",
                  "approved_at": _now(), "approved_by": user["id"],
                  "updated_at": _now()}},
    )
    await audit_log(user_id=user["id"], entity_type="service_charge",
                    entity_id=sc_id, action="approve")
    return await get_service_charge(sc_id)


async def post_service_charge(sc_id: str, *, user: dict) -> dict:
    db = get_db()
    d = await db.service_charge_periods.find_one({"id": sc_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Service charge tidak ditemukan")
    if d["status"] not in ("approved", "calculated"):
        raise ValidationError(f"Status saat ini: {d['status']}")
    je = await journal_service.post_for_service_charge(d, user_id=user["id"])
    # LB Fund ledger entry (in)
    if d.get("lb_amount", 0) > 0:
        await _lb_ledger_add(
            entry_date=_today(),
            direction="in",
            amount=float(d["lb_amount"]),
            source_type="service_charge",
            source_id=sc_id,
            outlet_id=d.get("outlet_id"),
            description=f"L&B deduction from service charge {d.get('period')}",
        )
    update: dict = {
        "status": "posted",
        "posted_at": _now(), "posted_by": user["id"],
        "journal_entry_id": je["id"] if je else None,
        "journal_skipped": je is None,
        "journal_skip_reason": (
            None if je else "Service charge total is 0 (no validated daily_sales for period+outlet)"
        ),
        "updated_at": _now(),
    }
    await db.service_charge_periods.update_one({"id": sc_id}, {"$set": update})
    await audit_log(user_id=user["id"], entity_type="service_charge",
                    entity_id=sc_id, action="post")
    return await get_service_charge(sc_id)


# ============================================================
# INCENTIVE
# ============================================================

async def list_schemes(*, page: int = 1, per_page: int = 50):
    db = get_db()
    q = {"deleted_at": None}
    skip = (page - 1) * per_page
    items = await db.incentive_schemes.find(q).sort([("created_at", -1)]).skip(skip).limit(per_page).to_list(per_page)
    total = await db.incentive_schemes.count_documents(q)
    return [serialize(d) for d in items], {"page": page, "per_page": per_page, "total": total}


async def create_scheme(payload: dict, *, user: dict) -> dict:
    db = get_db()
    code = (payload.get("code") or "").strip().upper()
    name = (payload.get("name") or "").strip()
    if not code or not name:
        raise ValidationError("code dan name wajib")
    if await db.incentive_schemes.find_one({"code": code, "deleted_at": None}):
        raise ConflictError(f"Scheme code {code} sudah ada")
    doc = {
        "id": str(uuid.uuid4()),
        "code": code,
        "name": name,
        "scope_type": payload.get("scope_type", "outlet"),
        "scope_id": payload.get("scope_id"),
        "rule_type": payload.get("rule_type", "pct_of_sales"),
        "rule_data": payload.get("rule_data") or {},
        "employee_ids": payload.get("employee_ids") or [],
        "active": True,
        "notes": payload.get("notes"),
        "created_at": _now(), "updated_at": _now(), "deleted_at": None,
        "created_by": user["id"],
    }
    await db.incentive_schemes.insert_one(doc)
    await audit_log(user_id=user["id"], entity_type="incentive_scheme",
                    entity_id=doc["id"], action="create")
    return serialize(doc)


async def list_runs(*, scheme_id: Optional[str] = None, period: Optional[str] = None,
                     status: Optional[str] = None, page: int = 1, per_page: int = 20):
    db = get_db()
    q: dict = {"deleted_at": None}
    if scheme_id:
        q["scheme_id"] = scheme_id
    if period:
        q["period"] = period
    if status:
        q["status"] = status
    skip = (page - 1) * per_page
    items = await db.incentive_runs.find(q).sort([("period", -1), ("created_at", -1)]).skip(skip).limit(per_page).to_list(per_page)
    total = await db.incentive_runs.count_documents(q)
    return [serialize(d) for d in items], {"page": page, "per_page": per_page, "total": total}


async def get_run(run_id: str) -> dict:
    db = get_db()
    d = await db.incentive_runs.find_one({"id": run_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Incentive run tidak ditemukan")
    return serialize(d)


async def calculate_incentive(payload: dict, *, user: dict) -> dict:
    """Run a scheme for a given period.
    Supported rule_types:
      - pct_of_sales: rule_data.pct (decimal). base = validated daily_sales for outlet+period.
      - flat_per_target: rule_data.target_sales, rule_data.flat_amount.
                          If base_sales >= target → flat_amount distributed.
      - tiered_sales: rule_data.tiers[{min_sales, max_sales, pct, flat}] → first match.
    Distribute equally across scheme.employee_ids.
    """
    db = get_db()
    scheme_id = payload.get("scheme_id")
    period = payload.get("period") or _period_now()
    if not scheme_id:
        raise ValidationError("scheme_id wajib")
    scheme = await db.incentive_schemes.find_one({"id": scheme_id, "deleted_at": None})
    if not scheme:
        raise ValidationError("Scheme tidak ditemukan")
    if not scheme.get("active"):
        raise ValidationError("Scheme tidak aktif")

    outlet_id = scheme.get("scope_id") if scheme.get("scope_type") == "outlet" else payload.get("outlet_id")
    # Compute base sales (validated daily_sales grand_total in period)
    sales_match: dict = {
        "deleted_at": None, "status": "validated",
        "sales_date": {"$gte": f"{period}-01", "$lte": f"{period}-31"},
    }
    if outlet_id:
        sales_match["outlet_id"] = outlet_id
    agg = db.daily_sales.aggregate([
        {"$match": sales_match},
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$grand_total", 0]}}}},
    ])
    res = await agg.to_list(1)
    base_sales = float(res[0]["total"]) if res else 0.0

    rule_type = scheme.get("rule_type", "pct_of_sales")
    rule_data = scheme.get("rule_data") or {}
    if rule_type == "pct_of_sales":
        pct = float(rule_data.get("pct", 0) or 0)
        total_amount = round(base_sales * pct, 2)
    elif rule_type == "flat_per_target":
        target = float(rule_data.get("target_sales", 0) or 0)
        flat = float(rule_data.get("flat_amount", 0) or 0)
        total_amount = flat if base_sales >= target else 0
    elif rule_type == "tiered_sales":
        total_amount = 0
        for tier in rule_data.get("tiers", []):
            mn = float(tier.get("min_sales", 0) or 0)
            mx = float(tier.get("max_sales", 9e18) or 9e18)
            if mn <= base_sales <= mx:
                total_amount = round(base_sales * float(tier.get("pct", 0) or 0)
                                     + float(tier.get("flat", 0) or 0), 2)
                break
    else:
        total_amount = 0

    # Distribute equally among employees in scheme
    emp_ids = scheme.get("employee_ids") or []
    allocations: list[dict] = []
    if emp_ids and total_amount > 0:
        per_head = round(total_amount / len(emp_ids), 2)
        last = round(total_amount - per_head * (len(emp_ids) - 1), 2)
        # Resolve names
        emp_map = {}
        async for e in db.employees.find({"id": {"$in": emp_ids}}):
            emp_map[e["id"]] = e.get("full_name", e["id"])
        for i, eid in enumerate(emp_ids):
            allocations.append({
                "employee_id": eid,
                "employee_name": emp_map.get(eid, eid),
                "base_amount": base_sales,
                "formula_detail": f"{rule_type}",
                "amount": last if i == len(emp_ids) - 1 else per_head,
            })

    # Upsert run for (scheme, period, outlet)
    existing = await db.incentive_runs.find_one({
        "scheme_id": scheme_id, "period": period,
        "outlet_id": outlet_id, "deleted_at": None,
    })
    common = {
        "scheme_id": scheme_id,
        "scheme_name": scheme.get("name"),
        "period": period,
        "outlet_id": outlet_id,
        "brand_id": payload.get("brand_id"),
        "base_sales": round(base_sales, 2),
        "allocations": allocations,
        "total_amount": round(total_amount, 2),
        "status": "calculated",
        "calculated_at": _now(), "calculated_by": user["id"],
        "updated_at": _now(),
        "notes": payload.get("notes"),
    }
    if existing:
        if existing["status"] == "posted":
            raise ConflictError("Run sudah posted; tidak bisa di-recalculate")
        await db.incentive_runs.update_one({"id": existing["id"]}, {"$set": common})
        run_id = existing["id"]
    else:
        doc_no = f"INC-{period}-{(scheme.get('code') or 'GEN')}"
        doc = {
            "id": str(uuid.uuid4()), "doc_no": doc_no,
            **common,
            "approved_at": None, "approved_by": None,
            "posted_at": None, "posted_by": None,
            "journal_entry_id": None,
            "created_at": _now(), "deleted_at": None,
            "created_by": user["id"],
        }
        await db.incentive_runs.insert_one(doc)
        run_id = doc["id"]
    await audit_log(user_id=user["id"], entity_type="incentive_run",
                    entity_id=run_id, action="calculate",
                    after={"scheme_id": scheme_id, "period": period})
    return await get_run(run_id)


async def approve_incentive(run_id: str, *, user: dict) -> dict:
    db = get_db()
    d = await db.incentive_runs.find_one({"id": run_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Incentive run tidak ditemukan")
    if d["status"] != "calculated":
        raise ValidationError(f"Status saat ini: {d['status']}")
    await db.incentive_runs.update_one(
        {"id": run_id},
        {"$set": {"status": "approved",
                  "approved_at": _now(), "approved_by": user["id"],
                  "updated_at": _now()}},
    )
    await audit_log(user_id=user["id"], entity_type="incentive_run",
                    entity_id=run_id, action="approve")
    return await get_run(run_id)


async def post_incentive(run_id: str, *, user: dict) -> dict:
    db = get_db()
    d = await db.incentive_runs.find_one({"id": run_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Incentive run tidak ditemukan")
    if d["status"] not in ("approved", "calculated"):
        raise ValidationError(f"Status saat ini: {d['status']}")
    je = await journal_service.post_for_incentive(d, user_id=user["id"])
    update: dict = {
        "status": "posted",
        "posted_at": _now(), "posted_by": user["id"],
        "journal_entry_id": je["id"] if je else None,
        "journal_skipped": je is None,
        "journal_skip_reason": (
            None if je else "Incentive total is 0 (no validated sales / scheme produced no payout)"
        ),
        "updated_at": _now(),
    }
    await db.incentive_runs.update_one({"id": run_id}, {"$set": update})
    await audit_log(user_id=user["id"], entity_type="incentive_run",
                    entity_id=run_id, action="post")
    return await get_run(run_id)


# ============================================================
# VOUCHER
# ============================================================

async def list_vouchers(
    *, status: Optional[str] = None, batch_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1, per_page: int = 50,
):
    db = get_db()
    q: dict = {"deleted_at": None}
    if status:
        q["status"] = status
    if batch_id:
        q["batch_id"] = batch_id
    if search:
        q["$or"] = [
            {"code": {"$regex": search, "$options": "i"}},
            {"purpose": {"$regex": search, "$options": "i"}},
            {"notes": {"$regex": search, "$options": "i"}},
        ]
    skip = (page - 1) * per_page
    items = await db.vouchers.find(q).sort([("issue_date", -1), ("created_at", -1)]).skip(skip).limit(per_page).to_list(per_page)
    total = await db.vouchers.count_documents(q)
    return [serialize(d) for d in items], {"page": page, "per_page": per_page, "total": total}


async def issue_vouchers(payload: dict, *, user: dict) -> dict:
    """Issue voucher batch:
    payload: {value, qty, expire_date?, purpose, prefix?, outlet_id?, post_journal: bool}
    """
    db = get_db()
    qty = int(payload.get("qty", 1) or 1)
    if qty < 1 or qty > 1000:
        raise ValidationError("qty 1..1000")
    value = float(payload.get("value", 0) or 0)
    if value <= 0:
        raise ValidationError("value harus > 0")
    purpose = (payload.get("purpose") or "marketing").strip().lower()
    if purpose not in ("marketing", "customer_comp", "staff", "replacement"):
        raise ValidationError("purpose tidak valid")
    prefix = (payload.get("prefix") or "VOC").upper()
    issue_date = payload.get("issue_date") or _today()
    expire_date = payload.get("expire_date")
    outlet_id = payload.get("outlet_id")
    post_je = bool(payload.get("post_journal", True))

    batch_id = str(uuid.uuid4())
    docs: list[dict] = []
    journals_created = 0
    for i in range(qty):
        seq_no = await next_doc_no("VOC")
        v = {
            "id": str(uuid.uuid4()),
            "code": f"{prefix}-{seq_no}",
            "batch_id": batch_id,
            "value": value,
            "issue_date": issue_date,
            "expire_date": expire_date,
            "issued_by": user["id"],
            "issued_to": payload.get("issued_to"),
            "purpose": purpose,
            "outlet_id": outlet_id,
            "status": "issued",
            "redeemed_at": None, "redeemed_amount": 0,
            "redeemed_outlet_id": None, "redeemed_ref": None,
            "journal_entry_issue_id": None, "journal_entry_redeem_id": None,
            "notes": payload.get("notes"),
            "created_at": _now(), "updated_at": _now(), "deleted_at": None,
            "created_by": user["id"],
        }
        await db.vouchers.insert_one(v)
        if post_je:
            je = await journal_service.post_for_voucher_issue(v, user_id=user["id"])
            if je:
                await db.vouchers.update_one({"id": v["id"]},
                    {"$set": {"journal_entry_issue_id": je["id"]}})
                v["journal_entry_issue_id"] = je["id"]
                journals_created += 1
        docs.append(v)
    await audit_log(user_id=user["id"], entity_type="voucher_batch",
                    entity_id=batch_id, action="issue",
                    after={"qty": qty, "value": value, "purpose": purpose})
    return {
        "batch_id": batch_id,
        "qty": qty, "value": value, "purpose": purpose,
        "vouchers": [serialize(v) for v in docs],
        "journals_created": journals_created,
    }


async def redeem_voucher(code: str, payload: dict, *, user: dict) -> dict:
    db = get_db()
    v = await db.vouchers.find_one({"code": code, "deleted_at": None})
    if not v:
        raise NotFoundError(f"Voucher {code} tidak ditemukan")
    if v["status"] != "issued":
        raise ValidationError(f"Voucher status: {v['status']}, tidak bisa redeem")
    if v.get("expire_date"):
        try:
            exp = datetime.strptime(v["expire_date"], "%Y-%m-%d").date()
            if exp < datetime.now(timezone.utc).date():
                await db.vouchers.update_one({"id": v["id"]},
                    {"$set": {"status": "expired", "updated_at": _now()}})
                raise ValidationError("Voucher sudah expired")
        except ValueError:
            pass
    redeemed_amount = float(payload.get("amount") or v["value"])
    if redeemed_amount <= 0:
        raise ValidationError("Amount redeem harus > 0")
    if redeemed_amount > float(v["value"]):
        raise ValidationError("Amount redeem melebihi value voucher")
    await db.vouchers.update_one(
        {"id": v["id"]},
        {"$set": {"status": "redeemed",
                  "redeemed_at": _now(),
                  "redeemed_amount": redeemed_amount,
                  "redeemed_outlet_id": payload.get("outlet_id"),
                  "redeemed_ref": payload.get("ref"),
                  "updated_at": _now()}},
    )
    fresh = await db.vouchers.find_one({"id": v["id"]})
    je = await journal_service.post_for_voucher_redeem(fresh, user_id=user["id"])
    if je:
        await db.vouchers.update_one({"id": v["id"]},
            {"$set": {"journal_entry_redeem_id": je["id"]}})
        fresh["journal_entry_redeem_id"] = je["id"]
    await audit_log(user_id=user["id"], entity_type="voucher",
                    entity_id=v["id"], action="redeem",
                    after={"amount": redeemed_amount})
    return serialize(fresh)


# ============================================================
# FOC
# ============================================================

async def list_foc(
    *, outlet_ids: Optional[list[str]] = None, foc_type: Optional[str] = None,
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    page: int = 1, per_page: int = 20,
):
    db = get_db()
    q: dict = {"deleted_at": None}
    if outlet_ids:
        q["outlet_id"] = {"$in": outlet_ids}
    if foc_type:
        q["foc_type"] = foc_type
    if date_from:
        q.setdefault("foc_date", {})["$gte"] = date_from
    if date_to:
        q.setdefault("foc_date", {})["$lte"] = date_to
    skip = (page - 1) * per_page
    items = await db.foc_entries.find(q).sort([("foc_date", -1), ("created_at", -1)]).skip(skip).limit(per_page).to_list(per_page)
    total = await db.foc_entries.count_documents(q)
    return [serialize(d) for d in items], {"page": page, "per_page": per_page, "total": total}


async def create_foc(payload: dict, *, user: dict) -> dict:
    db = get_db()
    outlet_id = payload.get("outlet_id")
    if not outlet_id:
        raise ValidationError("outlet_id wajib")
    outlet = await db.outlets.find_one({"id": outlet_id, "deleted_at": None})
    if not outlet:
        raise ValidationError("Outlet tidak ditemukan")
    foc_type = (payload.get("foc_type") or "").lower()
    if foc_type not in ("staff_meal", "marketing", "customer_comp", "other"):
        raise ValidationError("foc_type harus staff_meal/marketing/customer_comp/other")
    amount = float(payload.get("amount", 0) or 0)
    items = payload.get("items") or []
    if amount <= 0 and items:
        amount = sum(float(it.get("total", 0) or 0) for it in items)
    if amount <= 0:
        raise ValidationError("Amount > 0 wajib (atau items dengan total)")

    doc_no = await next_doc_no("FOC")
    doc = {
        "id": str(uuid.uuid4()),
        "doc_no": doc_no,
        "foc_date": payload.get("foc_date") or _today(),
        "outlet_id": outlet_id,
        "brand_id": outlet.get("brand_id"),
        "foc_type": foc_type,
        "amount": round(amount, 2),
        "items": items,
        "beneficiary": payload.get("beneficiary"),
        "gl_account_id": payload.get("gl_account_id"),
        "receipt_url": payload.get("receipt_url"),
        "notes": payload.get("notes"),
        "status": "posted",
        "journal_entry_id": None,
        "created_at": _now(), "updated_at": _now(), "deleted_at": None,
        "created_by": user["id"],
    }
    await db.foc_entries.insert_one(doc)
    je = await journal_service.post_for_foc(doc, user_id=user["id"])
    if je:
        await db.foc_entries.update_one({"id": doc["id"]},
            {"$set": {"journal_entry_id": je["id"]}})
        doc["journal_entry_id"] = je["id"]
    # Customer comp also feeds LB Fund out (when LB pays the comp)
    if foc_type == "customer_comp":
        await _lb_ledger_add(
            entry_date=doc["foc_date"],
            direction="out",
            amount=float(doc["amount"]),
            source_type="customer_compensation",
            source_id=doc["id"],
            outlet_id=outlet_id,
            description=f"Customer compensation {doc_no}",
        )
    await audit_log(user_id=user["id"], entity_type="foc",
                    entity_id=doc["id"], action="create")
    return serialize(doc)


# ============================================================
# LB FUND LEDGER
# ============================================================

async def _lb_ledger_add(
    *, entry_date: str, direction: str, amount: float,
    source_type: str, source_id: Optional[str] = None,
    outlet_id: Optional[str] = None, description: Optional[str] = None,
) -> dict:
    db = get_db()
    # Compute new running balance
    cursor = db.lb_fund_ledger.aggregate([
        {"$match": {"deleted_at": None}},
        {"$group": {
            "_id": None,
            "balance": {"$sum": {"$cond": [
                {"$eq": ["$direction", "in"]}, "$amount", {"$multiply": ["$amount", -1]},
            ]}},
        }},
    ])
    res = await cursor.to_list(1)
    cur = float(res[0]["balance"]) if res else 0.0
    delta = amount if direction == "in" else -amount
    new_bal = round(cur + delta, 2)
    doc = {
        "id": str(uuid.uuid4()),
        "entry_date": entry_date,
        "direction": direction,
        "amount": round(amount, 2),
        "source_type": source_type,
        "source_id": source_id,
        "outlet_id": outlet_id,
        "description": description,
        "balance_after": new_bal,
        "created_at": _now(), "updated_at": _now(), "deleted_at": None,
    }
    await db.lb_fund_ledger.insert_one(doc)
    return serialize(doc)


async def list_lb_fund(*, page: int = 1, per_page: int = 50):
    db = get_db()
    q = {"deleted_at": None}
    skip = (page - 1) * per_page
    items = await db.lb_fund_ledger.find(q).sort([("entry_date", -1), ("created_at", -1)]).skip(skip).limit(per_page).to_list(per_page)
    total = await db.lb_fund_ledger.count_documents(q)
    cursor = db.lb_fund_ledger.aggregate([
        {"$match": q},
        {"$group": {"_id": None,
                    "balance": {"$sum": {"$cond": [
                        {"$eq": ["$direction", "in"]}, "$amount", {"$multiply": ["$amount", -1]},
                    ]}}}},
    ])
    res = await cursor.to_list(1)
    balance = float(res[0]["balance"]) if res else 0.0
    return [serialize(d) for d in items], {"page": page, "per_page": per_page, "total": total, "balance": round(balance, 2)}


# ============================================================
# PAYROLL
# ============================================================

async def list_payroll(*, period: Optional[str] = None, status: Optional[str] = None,
                        page: int = 1, per_page: int = 20):
    db = get_db()
    q: dict = {"deleted_at": None}
    if period:
        q["period"] = period
    if status:
        q["status"] = status
    skip = (page - 1) * per_page
    items = await db.payroll_cycles.find(q).sort([("period", -1), ("created_at", -1)]).skip(skip).limit(per_page).to_list(per_page)
    total = await db.payroll_cycles.count_documents(q)
    return [serialize(d) for d in items], {"page": page, "per_page": per_page, "total": total}


async def get_payroll(p_id: str) -> dict:
    db = get_db()
    d = await db.payroll_cycles.find_one({"id": p_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Payroll cycle tidak ditemukan")
    return serialize(d)


async def create_payroll(payload: dict, *, user: dict) -> dict:
    """Generate payroll cycle (Sprint G enhanced):
    - Pull active employees (optionally filter by outlet)
    - For each: gross = basic + allowances + service_share + incentive
    - Deduct: BPJS employee share + PPh21 (if enabled) + advance_repayment
    - Take home = gross - deductions - advance_repay
    """
    from services.system_settings_service import get_value as _get_val  # noqa: avoid circular
    db = get_db()
    period = payload.get("period") or _period_now()
    outlet_id = payload.get("outlet_id")
    # Already-existing in-progress cycle for period+outlet?
    if await db.payroll_cycles.find_one({
        "period": period, "outlet_id": outlet_id,
        "status": {"$in": ["draft", "approved"]},
        "deleted_at": None,
    }):
        raise ConflictError(f"Payroll {period} sudah ada (draft/approved); selesaikan atau hapus dulu")

    # Tax feature flags
    pph21_enabled = str(await _get_val("TAX_PPH21_ENABLED") or "false").lower() == "true"
    pph21_method  = str(await _get_val("TAX_PPH21_METHOD") or "gross")

    emp_filter = {"deleted_at": None, "status": "active"}
    if outlet_id:
        emp_filter["outlet_id"] = outlet_id

    employees: list[dict] = []
    total_gross = 0.0
    total_bpjs_employee = 0.0
    total_pph21 = 0.0
    total_deductions = 0.0
    total_allowances = 0.0
    total_advance_repay = 0.0
    total_take_home = 0.0
    total_bpjs_employer = 0.0

    async for e in db.employees.find(emp_filter):
        # Salary master (Sprint G)
        sm = await db.salary_masters.find_one({"employee_id": e["id"], "deleted_at": None})

        basic = float(e.get("basic_salary", 0) or 0)
        if sm and sm.get("basic_salary"):
            basic = float(sm.get("basic_salary") or basic)

        # Allowances from salary master components
        allowances_total = 0.0
        allowances_list = []
        if sm:
            for comp in sm.get("components", []):
                amt = float(comp.get("amount", 0) or 0)
                allowances_total += amt
                allowances_list.append({
                    "code": comp.get("code", ""),
                    "name": comp.get("name", ""),
                    "amount": amt,
                })

        # Compute advance repay due for this period
        advances = await db.employee_advances.find({
            "employee_id": e["id"],
            "status": {"$in": ["repaying"]},
            "deleted_at": None,
        }).to_list(20)
        repay_amount = 0.0
        for a in advances:
            for line in a.get("schedule", []):
                if line.get("period") == period and not line.get("paid"):
                    repay_amount += float(line.get("amount", 0) or 0)

        # Service charge share
        sc_share = 0.0
        sc = await db.service_charge_periods.find_one({
            "period": period, "outlet_id": e.get("outlet_id"),
            "status": "posted", "deleted_at": None,
        })
        if sc:
            for alloc in sc.get("allocations", []):
                if alloc.get("employee_id") == e["id"]:
                    sc_share = float(alloc.get("amount", 0) or 0)
                    break

        # Incentive share
        inc_share = 0.0
        runs = await db.incentive_runs.find({
            "period": period, "status": "posted", "deleted_at": None,
        }).to_list(50)
        for r in runs:
            for alloc in r.get("allocations", []):
                if alloc.get("employee_id") == e["id"]:
                    inc_share += float(alloc.get("amount", 0) or 0)

        gross_total = round(basic + allowances_total + sc_share + inc_share, 2)

        # ── BPJS calculation ─────────────────────────────
        bpjs_enrolled = sm.get("bpjs_enrolled", True) if sm else True
        bpjs = _calc_bpjs(basic + allowances_total, bpjs_enrolled)  # BPJS base = fixed pay only

        # ── PPh 21 calculation ────────────────────────────
        pph21_monthly = 0.0
        pph21_detail: dict = {}
        if pph21_enabled:
            ptkp_status = sm.get("ptkp_status", "TK/0") if sm else "TK/0"
            taxable_monthly = gross_total  # include service/incentive in annualized estimate
            result21 = _calc_pph21(taxable_monthly, ptkp_status)
            pph21_monthly = result21["monthly_tax"]
            pph21_detail = result21
            pph21_detail["method"] = pph21_method

        deductions = round(bpjs["employee"] + pph21_monthly, 2)
        allowances_var = round(sc_share + inc_share, 2)  # variable part
        take_home = round(gross_total - deductions - repay_amount, 2)

        employees.append({
            "employee_id": e["id"],
            "name": e.get("full_name"),
            "outlet_id": e.get("outlet_id"),
            "basic": basic,
            "allowances_total": allowances_total,
            "allowances": allowances_list,
            "service_share": sc_share,
            "incentive_share": inc_share,
            "gross": gross_total,
            "bpjs_employee": bpjs["employee"],
            "bpjs_employer": bpjs["employer"],
            "bpjs_detail": bpjs.get("detail", {}),
            "pph21": pph21_monthly,
            "pph21_detail": pph21_detail,
            "deductions": deductions,
            "variable_pay": allowances_var,
            "advance_repayment": repay_amount,
            "take_home": take_home,
            "ptkp_status": sm.get("ptkp_status", "TK/0") if sm else "TK/0",
        })
        total_gross += gross_total
        total_bpjs_employee += bpjs["employee"]
        total_bpjs_employer += bpjs["employer"]
        total_pph21 += pph21_monthly
        total_deductions += deductions
        total_allowances += allowances_var
        total_advance_repay += repay_amount
        total_take_home += take_home

    doc_no = await next_doc_no("PAY")
    doc = {
        "id": str(uuid.uuid4()),
        "doc_no": doc_no,
        "period": period,
        "outlet_id": outlet_id,
        "payroll_date": payload.get("payroll_date") or f"{period}-25",
        "employees": employees,
        "total_gross": round(total_gross, 2),
        "total_deductions": round(total_deductions, 2),
        "total_allowances": round(total_allowances, 2),
        "total_bpjs_employee": round(total_bpjs_employee, 2),
        "total_bpjs_employer": round(total_bpjs_employer, 2),
        "total_pph21": round(total_pph21, 2),
        "total_advance_repayment": round(total_advance_repay, 2),
        "total_take_home": round(total_take_home, 2),
        "pph21_enabled": pph21_enabled,
        "status": "draft",
        "approved_at": None, "approved_by": None,
        "posted_at": None, "posted_by": None,
        "journal_entry_id": None,
        "notes": payload.get("notes"),
        "created_at": _now(), "updated_at": _now(), "deleted_at": None,
        "created_by": user["id"],
    }
    await db.payroll_cycles.insert_one(doc)
    await audit_log(user_id=user["id"], entity_type="payroll_cycle",
                    entity_id=doc["id"], action="create")
    return serialize(doc)


async def approve_payroll(p_id: str, *, user: dict) -> dict:
    db = get_db()
    d = await db.payroll_cycles.find_one({"id": p_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Payroll cycle tidak ditemukan")
    if d["status"] != "draft":
        raise ValidationError(f"Status saat ini: {d['status']}")
    await db.payroll_cycles.update_one(
        {"id": p_id},
        {"$set": {"status": "approved",
                  "approved_at": _now(), "approved_by": user["id"],
                  "updated_at": _now()}},
    )
    return await get_payroll(p_id)


async def post_payroll(p_id: str, *, user: dict) -> dict:
    db = get_db()
    d = await db.payroll_cycles.find_one({"id": p_id, "deleted_at": None})
    if not d:
        raise NotFoundError("Payroll cycle tidak ditemukan")
    if d["status"] not in ("approved", "draft"):
        raise ValidationError(f"Status saat ini: {d['status']}")
    je = await journal_service.post_for_payroll(d, user_id=user["id"])
    # Mark advance schedule lines paid for this period
    period = d.get("period")
    for emp in d.get("employees", []):
        if float(emp.get("advance_repayment", 0) or 0) > 0:
            advances = await db.employee_advances.find({
                "employee_id": emp["employee_id"],
                "status": "repaying",
                "deleted_at": None,
            }).to_list(20)
            for a in advances:
                schedule = a.get("schedule", [])
                changed = False
                for line in schedule:
                    if line.get("period") == period and not line.get("paid"):
                        line["paid"] = True
                        line["paid_at"] = _now()
                        changed = True
                        break
                if changed:
                    all_paid = all(item.get("paid") for item in schedule)
                    update: dict = {"schedule": schedule, "updated_at": _now()}
                    if all_paid:
                        update["status"] = "settled"
                        update["settled_at"] = _now()
                    await db.employee_advances.update_one(
                        {"id": a["id"]}, {"$set": update},
                    )
    await db.payroll_cycles.update_one(
        {"id": p_id},
        {"$set": {"status": "posted",
                  "posted_at": _now(), "posted_by": user["id"],
                  "journal_entry_id": je["id"] if je else None,
                  "updated_at": _now()}},
    )
    await audit_log(user_id=user["id"], entity_type="payroll_cycle",
                    entity_id=p_id, action="post")
    return await get_payroll(p_id)


# ============================================================
# DASHBOARD
# ============================================================

async def hr_dashboard() -> dict:
    db = get_db()
    period = _period_now()
    active_emp = await db.employees.count_documents({"deleted_at": None, "status": "active"})
    # Open advances (status repaying)
    open_advances = await db.employee_advances.count_documents({
        "deleted_at": None, "status": "repaying",
    })
    # Pending advance approval (draft)
    pending_adv = await db.employee_advances.count_documents({
        "deleted_at": None, "status": "draft",
    })
    # Total outstanding advance balance
    advances = await db.employee_advances.find({
        "deleted_at": None, "status": "repaying",
    }).to_list(1000)
    outstanding = 0.0
    for a in advances:
        for line in a.get("schedule", []):
            if not line.get("paid"):
                outstanding += float(line.get("amount", 0) or 0)
    # Service charge to allocate (calculated/approved status this period)
    sc_pending = await db.service_charge_periods.count_documents({
        "deleted_at": None, "period": period,
        "status": {"$in": ["calculated", "approved"]},
    })
    # Voucher stats
    issued_unredeemed = await db.vouchers.count_documents({
        "deleted_at": None, "status": "issued",
    })
    voucher_liab_cursor = db.vouchers.aggregate([
        {"$match": {"deleted_at": None, "status": "issued"}},
        {"$group": {"_id": None, "total": {"$sum": "$value"}}},
    ])
    res = await voucher_liab_cursor.to_list(1)
    voucher_liability = float(res[0]["total"]) if res else 0.0
    # LB fund balance
    cursor = db.lb_fund_ledger.aggregate([
        {"$match": {"deleted_at": None}},
        {"$group": {"_id": None,
                    "balance": {"$sum": {"$cond": [
                        {"$eq": ["$direction", "in"]}, "$amount", {"$multiply": ["$amount", -1]},
                    ]}}}},
    ])
    res2 = await cursor.to_list(1)
    lb_balance = float(res2[0]["balance"]) if res2 else 0.0
    # Incentive runs pending posting this period
    inc_pending = await db.incentive_runs.count_documents({
        "deleted_at": None, "period": period,
        "status": {"$in": ["calculated", "approved"]},
    })
    return {
        "period": period,
        "active_employees": active_emp,
        "open_advances": open_advances,
        "pending_advance_approval": pending_adv,
        "advance_outstanding": round(outstanding, 2),
        "service_charge_pending": sc_pending,
        "incentive_pending": inc_pending,
        "voucher_unredeemed_count": issued_unredeemed,
        "voucher_liability": round(voucher_liability, 2),
        "lb_fund_balance": round(lb_balance, 2),
    }



# ============================================================
# SALARY MASTER (Sprint G)
# ============================================================

STANDARD_COMPONENTS = [
    {"code": "TUNJ_JABATAN",   "name": "Tunjangan Jabatan",    "amount": 0.0},
    {"code": "TUNJ_MAKAN",     "name": "Tunjangan Makan",      "amount": 0.0},
    {"code": "TUNJ_TRANSPORT", "name": "Tunjangan Transport",  "amount": 0.0},
    {"code": "TUNJ_KESEHATAN", "name": "Tunjangan Kesehatan",  "amount": 0.0},
]

PTKP_OPTIONS = list(PTKP_BY_STATUS.keys())


async def get_salary_master(employee_id: str) -> dict | None:
    """Get salary master for an employee (or None if not set)."""
    db = get_db()
    emp = await db.employees.find_one({"id": employee_id, "deleted_at": None})
    if not emp:
        raise NotFoundError("Karyawan tidak ditemukan")
    sm = await db.salary_masters.find_one({"employee_id": employee_id, "deleted_at": None})
    if not sm:
        # Return default based on employee record
        basic = float(emp.get("basic_salary", 0) or 0)
        return {
            "employee_id": employee_id,
            "employee_name": emp.get("full_name"),
            "outlet_id": emp.get("outlet_id"),
            "basic_salary": basic,
            "components": STANDARD_COMPONENTS.copy(),
            "bpjs_enrolled": True,
            "ptkp_status": "TK/0",
            "npwp": emp.get("npwp", ""),
            "notes": "",
            "is_default": True,
        }
    return serialize({
        **sm,
        "employee_name": emp.get("full_name"),
        "outlet_id": emp.get("outlet_id"),
        "is_default": False,
    })


async def set_salary_master(employee_id: str, payload: dict, *, user: dict) -> dict:
    """Create or update salary master for an employee."""
    db = get_db()
    emp = await db.employees.find_one({"id": employee_id, "deleted_at": None})
    if not emp:
        raise NotFoundError("Karyawan tidak ditemukan")
    now = _now()
    components = payload.get("components", [])
    # Normalize
    cleaned = []
    for c in components:
        amt = float(c.get("amount", 0) or 0)
        cleaned.append({
            "code": c.get("code", "MISC"),
            "name": c.get("name", ""),
            "amount": amt,
        })
    doc = {
        "basic_salary": float(payload.get("basic_salary", emp.get("basic_salary", 0)) or 0),
        "components": cleaned,
        "bpjs_enrolled": bool(payload.get("bpjs_enrolled", True)),
        "ptkp_status": payload.get("ptkp_status", "TK/0"),
        "npwp": (payload.get("npwp") or emp.get("npwp", "")).strip(),
        "notes": payload.get("notes", ""),
        "updated_at": now,
        "updated_by": user["id"],
    }
    existing = await db.salary_masters.find_one({"employee_id": employee_id, "deleted_at": None})
    if existing:
        await db.salary_masters.update_one({"id": existing["id"]}, {"$set": doc})
        updated = await db.salary_masters.find_one({"id": existing["id"]})
        return serialize(updated)
    else:
        doc["id"] = str(uuid.uuid4())
        doc["employee_id"] = employee_id
        doc["created_at"] = now
        doc["created_by"] = user["id"]
        doc["deleted_at"] = None
        await db.salary_masters.insert_one(doc)
        return serialize(doc)


async def list_salary_masters(outlet_id: str | None = None, per_page: int = 100) -> list:
    """List salary masters with employee info."""
    db = get_db()
    emp_filter: dict = {"deleted_at": None, "status": "active"}
    if outlet_id:
        emp_filter["outlet_id"] = outlet_id
    emps = await db.employees.find(emp_filter).to_list(per_page)
    result = []
    for emp in emps:
        sm = await db.salary_masters.find_one({"employee_id": emp["id"], "deleted_at": None})
        basic = float(emp.get("basic_salary", 0) or 0)
        if sm:
            basic = float(sm.get("basic_salary", basic) or basic)
        allowances_total = sum(float(c.get("amount", 0) or 0) for c in (sm or {}).get("components", []))
        bpjs = _calc_bpjs(basic + allowances_total, (sm or {}).get("bpjs_enrolled", True))
        pph21_detail = _calc_pph21(basic + allowances_total, (sm or {}).get("ptkp_status", "TK/0"))
        result.append({
            "employee_id": emp["id"],
            "employee_name": emp.get("full_name"),
            "employee_code": emp.get("employee_id"),
            "outlet_id": emp.get("outlet_id"),
            "position": emp.get("position"),
            "basic_salary": basic,
            "allowances_total": allowances_total,
            "total_fixed_pay": round(basic + allowances_total, 2),
            "bpjs_enrolled": (sm or {}).get("bpjs_enrolled", True),
            "bpjs_employee": bpjs["employee"],
            "bpjs_employer": bpjs["employer"],
            "ptkp_status": (sm or {}).get("ptkp_status", "TK/0"),
            "pph21_monthly": pph21_detail["monthly_tax"],
            "npwp": (sm or {}).get("npwp") or emp.get("npwp", ""),
            "has_master": sm is not None,
        })
    return result


async def import_salary_excel(file_bytes: bytes, *, user: dict) -> dict:
    """Import salary master from Excel (.xlsx) or CSV.
    Expected columns: employee_code/employee_id, basic_salary, bpjs_enrolled (true/false),
                      ptkp_status, npwp, tunjangan_jabatan, tunjangan_makan, tunjangan_transport.
    Returns: {imported, updated, errors, preview}
    """
    import io as _io
    db = get_db()
    errors = []
    rows_in = []

    # Detect format
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes))
        ws = wb.active
        headers = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue
            rows_in.append({"_row": i, **dict(zip(headers, row))})
    except Exception:
        # Try CSV
        try:
            reader = csv.DictReader(_io.StringIO(file_bytes.decode("utf-8-sig")))
            for i, row in enumerate(reader, start=2):
                rows_in.append({"_row": i, **{k.strip().lower().replace(" ", "_"): v for k, v in row.items()}})
        except Exception as e:
            return {"imported": 0, "updated": 0, "errors": [f"Cannot parse file: {e}"], "preview": []}

    imported = 0
    updated = 0
    preview = []

    for row in rows_in:
        rn = row.get("_row", "?")
        # Identify employee
        emp = None
        for key in ("employee_code", "employee_id", "kode_karyawan", "nik"):
            val = str(row.get(key) or "").strip()
            if val:
                emp = await db.employees.find_one({
                    "$or": [{"employee_id": val}, {"id": val}, {"npwp": val}],
                    "deleted_at": None,
                })
                if emp:
                    break
        if not emp:
            # Try by name
            name_val = str(row.get("full_name") or row.get("nama") or "").strip()
            if name_val:
                emp = await db.employees.find_one({"full_name": {"$regex": name_val, "$options": "i"}, "deleted_at": None})
        if not emp:
            errors.append(f"Row {rn}: Employee not found (code/name missing or invalid)")
            continue

        def _float(key, default=0.0):
            v = row.get(key)
            try:
                return float(str(v).replace(",", "").strip()) if v not in (None, "") else default
            except Exception:
                return default

        basic = _float("basic_salary") or _float("gaji_pokok") or float(emp.get("basic_salary", 0) or 0)
        components = []
        for code, name, col in [
            ("TUNJ_JABATAN",   "Tunjangan Jabatan",   "tunjangan_jabatan"),
            ("TUNJ_MAKAN",     "Tunjangan Makan",     "tunjangan_makan"),
            ("TUNJ_TRANSPORT", "Tunjangan Transport", "tunjangan_transport"),
            ("TUNJ_KESEHATAN", "Tunjangan Kesehatan", "tunjangan_kesehatan"),
        ]:
            amt = _float(col)
            components.append({"code": code, "name": name, "amount": amt})

        bpjs_val = str(row.get("bpjs_enrolled", "true") or "true").lower().strip()
        bpjs_enrolled = bpjs_val not in ("false", "0", "no", "tidak")
        ptkp_status = str(row.get("ptkp_status") or "TK/0").strip().upper()
        if ptkp_status not in PTKP_BY_STATUS:
            ptkp_status = "TK/0"
        npwp = str(row.get("npwp") or emp.get("npwp", "") or "").strip()

        existing = await db.salary_masters.find_one({"employee_id": emp["id"], "deleted_at": None})
        now = _now()
        if existing:
            await db.salary_masters.update_one({"id": existing["id"]}, {"$set": {
                "basic_salary": basic, "components": components,
                "bpjs_enrolled": bpjs_enrolled, "ptkp_status": ptkp_status,
                "npwp": npwp, "updated_at": now, "updated_by": user["id"],
            }})
            updated += 1
        else:
            await db.salary_masters.insert_one({
                "id": str(uuid.uuid4()), "employee_id": emp["id"],
                "basic_salary": basic, "components": components,
                "bpjs_enrolled": bpjs_enrolled, "ptkp_status": ptkp_status,
                "npwp": npwp, "notes": "",
                "created_at": now, "created_by": user["id"],
                "updated_at": now, "updated_by": user["id"],
                "deleted_at": None,
            })
            imported += 1

        allowances_total = sum(c["amount"] for c in components)
        preview.append({
            "employee_name": emp.get("full_name"),
            "employee_id": emp["id"],
            "basic_salary": basic,
            "allowances_total": allowances_total,
            "ptkp_status": ptkp_status,
            "bpjs_enrolled": bpjs_enrolled,
        })

    return {
        "imported": imported,
        "updated": updated,
        "errors": errors,
        "preview": preview,
    }


async def get_payroll_payslip_data(cycle_id: str, employee_id: str) -> dict:
    """Return payslip data for one employee in a payroll cycle."""
    db = get_db()
    cycle = await db.payroll_cycles.find_one({"id": cycle_id, "deleted_at": None})
    if not cycle:
        raise NotFoundError("Payroll cycle tidak ditemukan")
    emp_data = next((e for e in cycle.get("employees", []) if e.get("employee_id") == employee_id), None)
    if not emp_data:
        raise NotFoundError("Karyawan tidak ada dalam payroll cycle ini")
    # Company info
    from services.system_settings_service import get_value as _get_val
    company_name = await _get_val("COMPANY_PKP_NAME") or "PT. Torado Group"
    company_npwp = await _get_val("COMPANY_NPWP") or ""
    return {
        "cycle_id": cycle_id,
        "doc_no": cycle.get("doc_no"),
        "period": cycle.get("period"),
        "payroll_date": cycle.get("payroll_date"),
        "company_name": company_name,
        "company_npwp": company_npwp,
        "employee": emp_data,
        "pph21_enabled": cycle.get("pph21_enabled", False),
    }
