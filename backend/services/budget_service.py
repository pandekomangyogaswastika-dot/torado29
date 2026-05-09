"""Budget vs Actual service — Sprint 2."""
from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Optional

from core.db import get_db, serialize
from models.budget import make_budget_doc, BUDGET_CATEGORIES

logger = logging.getLogger("aurora.budget")

# COA type-to-category mapping
COA_CATEGORY_MAP = {
    "4": "REV",   # revenue accounts (code starts with 4)
    "5": "COGS",  # COGS (code starts with 5)
    "6": "DEP",   # depreciation
    "50": "COGS",
    "51": "COGS",
    "52": "OPEX",
    "54": "PAYROLL",
    "55": "MKTG",
    "56": "OPEX",
    "57": "DEP",
    "58": "TAX",
    "59": "OPEX",
}


def guess_category(coa_code: str) -> str:
    code = str(coa_code)
    if code.startswith("4"):
        return "REV"
    if code.startswith("54"):
        return "PAYROLL"
    if code.startswith("55"):
        return "MKTG"
    if code.startswith("57") or code.startswith("61"):
        return "DEP"
    if code.startswith("58"):
        return "TAX"
    if code.startswith("5"):
        return "COGS"
    if code.startswith("6"):
        return "OPEX"
    return "OPEX"


async def create_budget(payload: dict, *, user_id: str) -> dict:
    db = get_db()
    # Enrich lines with COA info
    lines = []
    for line in payload.get("lines", []):
        coa = await db.chart_of_accounts.find_one({"id": line["coa_id"], "deleted_at": None})
        lines.append({
            "coa_id": line["coa_id"],
            "coa_code": (coa or {}).get("code"),
            "coa_name": (coa or {}).get("name"),
            "category": line.get("category") or guess_category((coa or {}).get("code", "")),
            "amount": round(float(line.get("amount", 0)), 2),
        })
    doc = make_budget_doc(
        name=payload.get("name") or f"Budget {payload.get('period')}",
        period=payload["period"],
        period_type=payload.get("period_type", "monthly"),
        outlet_id=payload.get("outlet_id"),
        lines=lines,
        notes=payload.get("notes"),
        created_by=user_id,
    )
    await db.budgets.insert_one(doc)
    return serialize(doc)


async def list_budgets(period: Optional[str] = None, outlet_id: Optional[str] = None) -> list[dict]:
    db = get_db()
    q: dict = {"deleted_at": None, "status": "active"}
    if period:
        q["period"] = period
    if outlet_id:
        q["outlet_id"] = outlet_id
    items = await db.budgets.find(q).sort([("period", -1)]).to_list(100)
    return [serialize(i) for i in items]


async def get_budget(budget_id: str) -> Optional[dict]:
    db = get_db()
    doc = await db.budgets.find_one({"id": budget_id, "deleted_at": None})
    return serialize(doc) if doc else None


async def update_budget(budget_id: str, payload: dict, *, user_id: str) -> Optional[dict]:
    db = get_db()
    lines = []
    if "lines" in payload:
        for line in payload["lines"]:
            coa = await db.chart_of_accounts.find_one({"id": line["coa_id"], "deleted_at": None})
            lines.append({
                "coa_id": line["coa_id"],
                "coa_code": (coa or {}).get("code"),
                "coa_name": (coa or {}).get("name"),
                "category": line.get("category") or guess_category((coa or {}).get("code", "")),
                "amount": round(float(line.get("amount", 0)), 2),
            })
        await db.budgets.update_one(
            {"id": budget_id},
            {"$set": {"lines": lines, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    return await get_budget(budget_id)


async def delete_budget(budget_id: str) -> bool:
    db = get_db()
    await db.budgets.update_one({"id": budget_id}, {"$set": {"status": "archived", "deleted_at": datetime.now(timezone.utc).isoformat()}})
    return True


async def vs_actual(
    period: str,
    outlet_id: Optional[str] = None,
    level: str = "coa",  # coa | category | both
) -> dict:
    """Compute budget vs actual for a period.

    Actual = sum of JE debit/credit for each COA in posted JE for the period.
    Budget = from budgets collection.
    """
    db = get_db()

    # 1. Get budget
    q_bud: dict = {"period": period, "deleted_at": None, "status": "active"}
    if outlet_id:
        q_bud["outlet_id"] = outlet_id
    budget_docs = await db.budgets.find(q_bud).to_list(10)
    budget_by_coa: dict[str, float] = {}
    budget_by_cat: dict[str, float] = {}
    for bdoc in budget_docs:
        for line in bdoc.get("lines", []):
            coa_id = line["coa_id"]
            amount = float(line.get("amount", 0))
            budget_by_coa[coa_id] = budget_by_coa.get(coa_id, 0) + amount
            cat = line.get("category", "OPEX")
            budget_by_cat[cat] = budget_by_cat.get(cat, 0) + amount

    # 2. Get actuals from JE
    actual_by_coa: dict[str, float] = {}
    async for je in db.journal_entries.find({"period": period, "status": "posted", "deleted_at": None}):
        for line in je.get("lines", []):
            coa_id = line.get("coa_id")
            if not coa_id:
                continue
            # For P&L: revenue = credit-normal, expense = debit-normal
            net = float(line.get("dr", 0)) - float(line.get("cr", 0))
            actual_by_coa[coa_id] = actual_by_coa.get(coa_id, 0) + net

    # 3. Build COA-level comparison
    all_coa_ids = set(list(budget_by_coa.keys()) + list(actual_by_coa.keys()))
    coa_level = []
    for coa_id in all_coa_ids:
        coa_doc = await db.chart_of_accounts.find_one({"id": coa_id, "deleted_at": None})
        if not coa_doc:
            continue
        budget_amt = budget_by_coa.get(coa_id, 0)
        actual_amt = actual_by_coa.get(coa_id, 0)
        variance = actual_amt - budget_amt
        variance_pct = (variance / budget_amt * 100) if budget_amt != 0 else None
        cat = guess_category(coa_doc.get("code", ""))
        coa_level.append({
            "coa_id": coa_id,
            "coa_code": coa_doc.get("code"),
            "coa_name": coa_doc.get("name"),
            "category": cat,
            "budget": round(budget_amt, 2),
            "actual": round(actual_amt, 2),
            "variance": round(variance, 2),
            "variance_pct": round(variance_pct, 1) if variance_pct is not None else None,
            "status": "over" if variance > 0 and cat not in ("REV",) else "under" if variance < 0 else "on_target",
        })
    coa_level.sort(key=lambda x: x["coa_code"] or "")

    # 4. Category rollup
    cat_rollup: dict[str, dict] = {}
    for row in coa_level:
        cat = row["category"]
        if cat not in cat_rollup:
            cat_rollup[cat] = {"category": cat, "budget": 0, "actual": 0, "variance": 0}
        cat_rollup[cat]["budget"] += row["budget"]
        cat_rollup[cat]["actual"] += row["actual"]
    for cat, row in cat_rollup.items():
        row["variance"] = round(row["actual"] - row["budget"], 2)
        row["variance_pct"] = round(row["variance"] / row["budget"] * 100, 1) if row["budget"] != 0 else None
        row["budget"] = round(row["budget"], 2)
        row["actual"] = round(row["actual"], 2)

    # Sort category rollup by BUDGET_CATEGORIES order
    cat_order = [c["code"] for c in BUDGET_CATEGORIES if not c.get("derived")]
    cat_rows = sorted(
        cat_rollup.values(),
        key=lambda x: cat_order.index(x["category"]) if x["category"] in cat_order else 99
    )

    return {
        "period": period,
        "outlet_id": outlet_id,
        "coa_level": coa_level,
        "category_rollup": cat_rows,
        "total_budget": round(sum(r["budget"] for r in coa_level), 2),
        "total_actual": round(sum(r["actual"] for r in coa_level), 2),
        "total_variance": round(sum(r["variance"] for r in coa_level), 2),
    }


async def import_csv(csv_content: str, period: str, *, user_id: str) -> dict:
    """Import budget from CSV. Columns: coa_code, amount, category (optional)."""
    db = get_db()
    reader = csv.DictReader(io.StringIO(csv_content))
    lines = []
    errors = []
    for i, row in enumerate(reader, 1):
        coa_code = (row.get("coa_code") or row.get("code") or "").strip()
        amount_str = (row.get("amount") or "0").strip().replace(",", "")
        try:
            amount = float(amount_str)
        except ValueError:
            errors.append(f"Row {i}: invalid amount '{amount_str}'")
            continue
        coa = await db.chart_of_accounts.find_one({"code": coa_code, "deleted_at": None})
        if not coa:
            errors.append(f"Row {i}: COA code '{coa_code}' not found")
            continue
        lines.append({
            "coa_id": coa["id"],
            "coa_code": coa_code,
            "coa_name": coa["name"],
            "category": row.get("category") or guess_category(coa_code),
            "amount": amount,
        })
    if errors:
        return {"success": False, "errors": errors, "imported": 0}
    doc = make_budget_doc(
        name=f"Budget Import {period}",
        period=period,
        period_type="monthly",
        outlet_id=None,
        lines=lines,
        notes="Imported from CSV",
        created_by=user_id,
    )
    await db.budgets.insert_one(doc)
    return {"success": True, "imported": len(lines), "budget_id": doc["id"]}



async def import_excel(file_bytes: bytes, period: str, *, outlet_id: str | None = None, user_id: str) -> dict:
    """Import budget from Excel (.xlsx) file using openpyxl.
    Expected headers: coa_code, amount (required) + coa_name, category (optional).
    """
    import io as _io
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "errors": ["openpyxl not installed"], "imported": 0}

    db = get_db()
    lines = []
    errors = []

    try:
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes))
        ws = wb.active
        # Read headers from first row
        raw_headers = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue  # skip empty rows
            row_dict = dict(zip(raw_headers, row))
            coa_code = str(row_dict.get("coa_code") or row_dict.get("code") or "").strip()
            amount_raw = row_dict.get("amount") or row_dict.get("jumlah") or 0
            try:
                amount = float(str(amount_raw).replace(",", "").strip())
            except (ValueError, TypeError):
                errors.append(f"Row {i}: invalid amount '{amount_raw}'")
                continue
            if not coa_code:
                errors.append(f"Row {i}: missing coa_code")
                continue
            coa = await db.chart_of_accounts.find_one({"code": coa_code, "deleted_at": None})
            if not coa:
                errors.append(f"Row {i}: COA '{coa_code}' not found")
                continue
            category = str(row_dict.get("category") or row_dict.get("kategori") or "").strip() or guess_category(coa_code)
            lines.append({
                "coa_id": coa["id"],
                "coa_code": coa_code,
                "coa_name": coa["name"],
                "category": category,
                "amount": amount,
            })
    except Exception as e:
        return {"success": False, "errors": [f"Failed to read Excel: {e}"], "imported": 0}

    if not lines:
        return {"success": False, "errors": errors or ["No valid rows found"], "imported": 0}

    doc = make_budget_doc(
        name=f"Budget Import Excel {period}",
        period=period,
        period_type="monthly",
        outlet_id=outlet_id,
        lines=lines,
        notes=f"Imported from Excel ({len(lines)} rows)",
        created_by=user_id,
    )
    await db.budgets.insert_one(doc)
    return {
        "success": True,
        "imported": len(lines),
        "skipped_errors": len(errors),
        "errors": errors,
        "budget_id": doc["id"],
        "period": period,
    }
