"""TORADO ERP — Full Data Migration & Simulation Script.

Imports real Excel data + simulates missing data for a complete-looking system.

Source files:
- /app/excel_source/financial_report.xlsx (JAE 1840, PAY 1486, KB 832, ACC 180...)
- /app/excel_source/market_list.xlsx (MASTER 2114 items, per-brand availability)
- /app/excel_source/purchasing_report.xlsx (Master 12777 purchasing rows)

Phases:
  A. Master Data Refresh (COA, items, vendors, employees, payment methods)
  B. Real Transactional Import (JAE, KB, PAY, Purchasing, EA, Voucher, Tax)
  C. Smart Simulation (Daily Sales, Petty Cash, Customers, Loyalty, Movements...)
  D. Linking & Validation
"""
import asyncio
import os
import sys
import uuid
import logging
import random
from datetime import datetime, timezone, timedelta
from pathlib import Path
from collections import defaultdict

import openpyxl
from dotenv import load_dotenv
load_dotenv("/app/backend/.env")
from motor.motor_asyncio import AsyncIOMotorClient

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger("torado_migration")

# Excel sources
EXCEL_FIN = "/app/excel_source/financial_report.xlsx"
EXCEL_ML = "/app/excel_source/market_list.xlsx"
EXCEL_PUR = "/app/excel_source/purchasing_report.xlsx"

MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.getenv("DB_NAME", "aurora")

# Random seed for reproducibility
random.seed(42)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def to_iso(dt) -> str | None:
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.isoformat()
    try:
        return datetime.strptime(str(dt)[:10], "%Y-%m-%d").isoformat()
    except Exception:
        return None


def safe_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s in ("#REF!", "#N/A", "None", "nan"):
        return ""
    return s


def safe_num(v) -> float:
    if v is None:
        return 0.0
    try:
        s = str(v).replace(",", "").strip()
        if s in ("", "#REF!", "#N/A", "-"):
            return 0.0
        return float(s)
    except Exception:
        return 0.0


# ============================================================
# PHASE A — MASTER DATA REFRESH
# ============================================================

def categorize_account(name: str) -> tuple[str, str, str]:
    """Heuristic: classify account into (type, code_prefix, normal_balance)."""
    n = name.lower()
    # Asset
    if any(k in n for k in ["cash", "bank", "petty", "inventory", "ar ", "receivable",
                             "prepaid", "fixed asset", "equipment", "deposit", "vehicle",
                             "leasehold"]):
        return "asset", "1", "Dr"
    # Liability
    if any(k in n for k in ["payable", "ap ", "tax payable", "liability", "loan",
                             "owing", "accrued", "due to", "kontra", "salary payable",
                             "bpjs", "withholding", "utang"]):
        return "liability", "2", "Cr"
    # Equity
    if any(k in n for k in ["capital", "retained", "shu", "owner equity", "owner's equity"]):
        return "equity", "3", "Cr"
    # Revenue
    if any(k in n for k in ["sales", "revenue", "income", "service charge",
                             "banquete", "buffet sales", "voucher revenue", "discount"]):
        return "revenue", "4", "Cr"
    # COGS / Direct
    if any(k in n for k in ["cogs", "purchase", "kitchen materials", "bar materials",
                             "buffet materials", "floor (materials)", "consumption", "purchases"]):
        return "cogs", "5", "Dr"
    # Expense (default)
    return "expense", "6", "Dr"


async def phase_a_master(db):
    log.info("=" * 70)
    log.info("PHASE A — MASTER DATA REFRESH")
    log.info("=" * 70)

    # ---- A.1: Reset COA → reload 180 accounts from ACC sheet ----
    log.info("[A.1] Resetting Chart of Accounts → 180 accounts from Excel...")
    await db.chart_of_accounts.delete_many({})

    wb = openpyxl.load_workbook(EXCEL_FIN, data_only=True, read_only=True)
    ws = wb['ACC']
    raw_accounts = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None:
            n = safe_str(row[0])
            if n and n not in raw_accounts:
                raw_accounts.append(n)
    wb.close()

    coa_docs = []
    seq = defaultdict(int)
    coa_by_name = {}
    for name in raw_accounts:
        acc_type, prefix, normal = categorize_account(name)
        seq[prefix] += 1
        code = f"{prefix}{seq[prefix]:03d}"
        doc = {
            "id": str(uuid.uuid4()),
            "code": code,
            "name": name,
            "name_id": name,
            "type": acc_type,
            "level": 1,
            "normal_balance": normal,
            "is_postable": True,
            "active": True,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        coa_docs.append(doc)
        coa_by_name[name.lower()] = doc

    if coa_docs:
        await db.chart_of_accounts.insert_many(coa_docs)
        # IMPORTANT: insert_many mutates docs and adds _id; remove for safety
        for d in coa_docs:
            d.pop("_id", None)
    log.info(f"[A.1] Inserted {len(coa_docs)} COA accounts")

    # ---- A.2: Items from MASTER (no edit) ----
    log.info("[A.2] Re-importing items from Market List MASTER...")
    await db.items.delete_many({})
    await db.item_pricings.delete_many({})

    wb = openpyxl.load_workbook(EXCEL_ML, data_only=True, read_only=True)
    ws = wb['MASTER (no edit)']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Headers at row 3 (index 2): ID, Regist Date, Items, P_Q1_25, P_Q2_25, P_Q3_25, P_Q4_25, P_Q1_26, P_Q2_26, P_Q3_26, P_Q4_26, Unit(Prod), Unit(Cost), Convert, Category, Direct Purchase, Contra, Valid, Altero, MDS, Calluna, RP, GG, BK, E-Crew
    # Data starts row 4 (index 3)
    items_docs = []
    pricing_docs = []
    seen_codes = set()
    categories_set = set()

    # Get brand & outlet ids
    brands_by_code = {}
    async for b in db.brands.find({}):
        brands_by_code[b["code"].upper()] = b["id"]
    outlets_by_brand = {}
    async for o in db.outlets.find({}):
        outlets_by_brand[o["brand_id"]] = o["id"]

    PERIOD_DATES = [
        ("2025-01-01", "2025-03-31", "2025-Q1"),
        ("2025-04-01", "2025-06-30", "2025-Q2"),
        ("2025-07-01", "2025-09-30", "2025-Q3"),
        ("2025-10-01", "2025-12-31", "2025-Q4"),
        ("2026-01-01", "2026-03-31", "2026-Q1"),
        ("2026-04-01", "2026-06-30", "2026-Q2"),
        ("2026-07-01", "2026-09-30", "2026-Q3"),
        ("2026-10-01", "2026-12-31", "2026-Q4"),
    ]

    for r_idx, row in enumerate(rows[3:], start=4):
        code_raw = safe_str(row[0])
        item_name = safe_str(row[2])
        if not item_name:
            continue
        # Use sequential code if duplicate
        code = code_raw or f"ITM-{r_idx:05d}"
        if code in seen_codes:
            code = f"{code}-{r_idx}"
        seen_codes.add(code)

        unit_prod = safe_str(row[11]) or "pcs"
        unit_cost = safe_str(row[12]) or unit_prod
        convert_unit = safe_num(row[13]) or 1
        category = safe_str(row[14]) or "Uncategorized"
        is_direct = safe_str(row[15]).lower() == "true"
        # Brand availability flags row[18..23] = Altero, MDS, Calluna, RP, GG, BK
        # Map Excel codes → DB codes (DB has: ALT, DLS, CAL, RKP, BKK)
        brand_flags_excel = {
            "ALT": safe_str(row[18]).lower() == "true",     # Altero
            "DLS": safe_str(row[19]).lower() == "true",     # MDS = Maison de la Sol
            "CAL": safe_str(row[20]).lower() == "true",     # Calluna
            "RKP": safe_str(row[21]).lower() == "true",     # RP = Rucker Park
            "BKK": safe_str(row[23]).lower() == "true" if len(row) > 23 else False,  # BK = Bakkies
        }
        brand_availability = [brands_by_code[c] for c, v in brand_flags_excel.items() if v and c in brands_by_code]
        categories_set.add(category)

        # Latest price: pick most recent non-empty in Q1 2025..Q4 2026
        prices = [safe_num(row[3+i]) for i in range(8)]
        latest_price = next((p for p in reversed(prices) if p > 0), 0)

        item_id = str(uuid.uuid4())
        item_doc = {
            "id": item_id,
            "code": code,
            "name": item_name,
            "name_local": item_name,
            "category": category,
            "unit_default": unit_prod,
            "conversion_units": [{"unit": unit_cost, "factor": convert_unit}] if unit_cost != unit_prod else [],
            "is_direct_purchase": is_direct,
            "brand_availability": brand_availability,
            "active": True,
            "price": latest_price,
            "par_levels": {},
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        items_docs.append(item_doc)

        # Create pricing history records
        prev_p = None
        for i, (eff_from, eff_to, label) in enumerate(PERIOD_DATES):
            p = prices[i]
            if p > 0:
                variance = None
                if prev_p and prev_p > 0:
                    variance = ((p - prev_p) / prev_p) * 100
                pricing_docs.append({
                    "id": str(uuid.uuid4()),
                    "item_id": item_id,
                    "vendor_id": None,
                    "unit": unit_cost,
                    "price": p,
                    "effective_from": eff_from,
                    "effective_to": eff_to if i < len(PERIOD_DATES) - 1 else None,
                    "is_active": (i == len(PERIOD_DATES) - 1),  # only last is active
                    "previous_price": prev_p,
                    "variance": variance,
                    "label": label,
                    "created_at": now_iso(),
                    "updated_at": now_iso(),
                })
                prev_p = p

    # Bulk insert in batches
    BATCH = 500
    inserted = 0
    for i in range(0, len(items_docs), BATCH):
        batch = items_docs[i:i+BATCH]
        await db.items.insert_many(batch)
        for d in batch:
            d.pop("_id", None)
        inserted += len(batch)
    log.info(f"[A.2] Inserted {inserted} items")

    inserted_p = 0
    for i in range(0, len(pricing_docs), BATCH):
        batch = pricing_docs[i:i+BATCH]
        await db.item_pricings.insert_many(batch)
        for d in batch:
            d.pop("_id", None)
        inserted_p += len(batch)
    log.info(f"[A.2] Inserted {inserted_p} item pricing records (multi-period)")

    # Create categories from items
    await db.categories.delete_many({"type": "item"})
    cat_docs = []
    for c in sorted(categories_set):
        cat_docs.append({
            "id": str(uuid.uuid4()),
            "type": "item",
            "code": c[:8].upper().replace(" ", ""),
            "name": c,
            "active": True,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })
    if cat_docs:
        await db.categories.insert_many(cat_docs)
        for d in cat_docs:
            d.pop("_id", None)
    log.info(f"[A.2] Inserted {len(cat_docs)} item categories")

    # ---- A.3: Vendors — extract unique from KB + Purchasing Master ----
    log.info("[A.3] Extracting unique vendors from KB + Purchasing data...")
    vendors_set = set()
    wb = openpyxl.load_workbook(EXCEL_FIN, data_only=True, read_only=True)
    ws = wb['KB']
    for row in ws.iter_rows(min_row=4, values_only=True):
        v = safe_str(row[8]) if len(row) > 8 else ""
        if v and v not in ("#N/A", "#REF!"):
            vendors_set.add(v)
    wb.close()

    wb = openpyxl.load_workbook(EXCEL_PUR, data_only=True, read_only=True)
    ws = wb['Master']
    for row in ws.iter_rows(min_row=4, values_only=True):
        v = safe_str(row[8]) if len(row) > 8 else ""
        if v and v not in ("#N/A", "#REF!"):
            vendors_set.add(v)
    wb.close()
    
    log.info(f"[A.3] Found {len(vendors_set)} unique vendor names")

    # Reset & re-insert vendors
    await db.vendors.delete_many({})
    vendor_docs = []
    vendor_by_name = {}
    for i, v_name in enumerate(sorted(vendors_set), 1):
        # Generate Indonesian-realistic contact data
        npwp = f"01.{random.randint(100, 999)}.{random.randint(100, 999)}.{random.randint(1, 9)}-{random.randint(100, 999)}.{random.randint(100, 999)}"
        bank_choices = ["BCA", "Mandiri", "BNI", "BRI", "Permata"]
        v_doc = {
            "id": str(uuid.uuid4()),
            "code": f"VND-{i:04d}",
            "name": v_name,
            "vendor_type": random.choice(["Food Supplier", "Beverage Supplier", "Operational", "Equipment", "Service"]),
            "contact_name": f"Bapak {random.choice(['Andi', 'Budi', 'Cahyo', 'Doni', 'Eko', 'Faisal', 'Gunawan', 'Hadi'])}",
            "phone": f"08{random.randint(10, 99)}-{random.randint(1000, 9999)}-{random.randint(1000, 9999)}",
            "email": f"sales@{v_name.lower().replace(' ', '').replace('/', '')[:15]}.co.id",
            "address": f"Jl. {random.choice(['Sudirman', 'Thamrin', 'Gatot Subroto', 'Diponegoro', 'Veteran', 'Asia Afrika'])} No. {random.randint(1, 200)}, Jakarta",
            "npwp": npwp,
            "bank_account": {
                "bank": random.choice(bank_choices),
                "account": f"{random.randint(1000000000, 9999999999)}",
                "name": v_name,
            },
            "default_payment_terms_days": random.choice([7, 14, 30, 30, 30, 45, 60]),
            "default_payment_method": "transfer",
            "active": True,
            "notes": f"Migrated from Excel data on {datetime.now().strftime('%Y-%m-%d')}",
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        vendor_docs.append(v_doc)
        vendor_by_name[v_name.lower()] = v_doc

    if vendor_docs:
        for i in range(0, len(vendor_docs), BATCH):
            batch = vendor_docs[i:i+BATCH]
            await db.vendors.insert_many(batch)
            for d in batch:
                d.pop("_id", None)
    log.info(f"[A.3] Inserted {len(vendor_docs)} vendors")

    # ---- A.4: Employees from EA, Voucher, Travel Incentive ----
    log.info("[A.4] Extracting employees from EA + Voucher + Service sheets...")
    employee_set = set()
    wb = openpyxl.load_workbook(EXCEL_FIN, data_only=True, read_only=True)
    if 'EA' in wb.sheetnames:
        ws = wb['EA']
        for row in ws.iter_rows(min_row=4, values_only=True):
            n = safe_str(row[5]) if len(row) > 5 else ""
            if n:
                employee_set.add(n)
    if 'Voucher' in wb.sheetnames:
        ws = wb['Voucher']
        for row in ws.iter_rows(min_row=4, values_only=True):
            n = safe_str(row[4]) if len(row) > 4 else ""
            if n:
                employee_set.add(n)
    if 'Travel Incentive' in wb.sheetnames:
        ws = wb['Travel Incentive']
        for row in ws.iter_rows(min_row=4, values_only=True):
            n = safe_str(row[1]) if len(row) > 1 else ""
            if n:
                employee_set.add(n)
    wb.close()

    # Augment with simulated employees per outlet
    SIM_NAMES = [
        "Ahmad Rizki Pratama", "Bambang Setiawan", "Citra Dewi Anjani", "Dedi Mulyadi",
        "Eka Sari Puspita", "Fajar Nugroho", "Gita Maharani", "Hadi Wijaya", "Indra Kusuma",
        "Joko Susilo", "Kartika Sari", "Lina Wati", "Maman Suherman", "Nita Lestari",
        "Oka Pratama", "Putri Ayu Wandari", "Qori Hidayat", "Rina Marlina",
        "Surya Gunawan", "Tati Suryani", "Udin Wahyudi", "Vera Anggraini",
        "Wawan Setiyo", "Xenia Putri", "Yanto Hardiyanto", "Zaki Maulana"
    ]
    for n in SIM_NAMES:
        employee_set.add(n)

    await db.employees.delete_many({})
    outlet_list = list(outlets_by_brand.values())
    emp_docs = []
    POSITIONS = ["Kitchen Staff", "Bar Staff", "Server", "Cashier", "Outlet Manager",
                 "Sous Chef", "Head Chef", "Bartender", "Barista", "Steward", "Captain"]
    for i, name in enumerate(sorted(employee_set), 1):
        outlet_id = random.choice(outlet_list) if outlet_list else None
        # Get brand from outlet
        brand_id = None
        if outlet_id:
            for bid, oid in outlets_by_brand.items():
                if oid == outlet_id:
                    brand_id = bid
                    break
        emp_docs.append({
            "id": str(uuid.uuid4()),
            "code": f"EMP-{i:04d}",
            "full_name": name,
            "position": random.choice(POSITIONS),
            "department": random.choice(["Kitchen", "Bar", "Service", "Office", "Management"]),
            "outlet_id": outlet_id,
            "brand_id": brand_id,
            "status": "active",
            "join_date": (datetime(2024, 1, 1) + timedelta(days=random.randint(0, 700))).strftime("%Y-%m-%d"),
            "bank_account": {
                "bank": random.choice(["BCA", "Mandiri", "BNI", "BRI"]),
                "account": str(random.randint(1000000000, 9999999999)),
                "name": name,
            },
            "npwp": f"{random.randint(10, 99)}.{random.randint(100, 999)}.{random.randint(100, 999)}.{random.randint(1, 9)}-{random.randint(100, 999)}.{random.randint(100, 999)}",
            "gross_salary": random.choice([4500000, 5000000, 6000000, 7500000, 8000000, 10000000, 12000000]),
            "basic_salary": random.choice([4000000, 4500000, 5000000, 6500000, 7500000, 9000000]),
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })
    if emp_docs:
        for i in range(0, len(emp_docs), BATCH):
            batch = emp_docs[i:i+BATCH]
            await db.employees.insert_many(batch)
            for d in batch:
                d.pop("_id", None)
    log.info(f"[A.4] Inserted {len(emp_docs)} employees")

    # ---- A.5: Payment Methods & Bank Accounts ----
    log.info("[A.5] Setting up payment methods & bank accounts...")
    await db.payment_methods.delete_many({})
    await db.bank_accounts.delete_many({})

    BANK_ACCOUNTS_DATA = [
        ("BCA-22-MIURA", "BCA 22 - Miura", "BCA", "0152200022"),
        ("BCA-56-OUT", "BCA 56 - Torado (OUT)", "BCA", "0152200056"),
        ("BCA-73-IN", "BCA 73 - Torado (IN)", "BCA", "0152200073"),
        ("MANDIRI-MIURA", "Mandiri - CV. Miura", "Mandiri", "1234500001"),
        ("MANDIRI-MALISA", "Mandiri - CV. Malisa", "Mandiri", "1234500002"),
        ("BNI-MIURA", "BNI - CV. Miura", "BNI", "9876500001"),
        ("BRI-MIURA", "Bank BRI - CV. Miura", "BRI", "5550100001"),
        ("BTN-MIURA", "Bank BTN - Miura", "BTN", "8001000001"),
        ("PERMATA-MIURA", "Bank Permata - CV. Miura", "Permata", "7000010001"),
    ]
    bank_docs = []
    cash_coa = await db.chart_of_accounts.find_one({"type": "asset", "name": {"$regex": "bank", "$options": "i"}}, {"_id": 0})
    for code, name, bank, acc_num in BANK_ACCOUNTS_DATA:
        bank_docs.append({
            "id": str(uuid.uuid4()),
            "code": code,
            "name": name,
            "bank": bank,
            "account_number": acc_num,
            "currency": "IDR",
            "gl_account_id": cash_coa["id"] if cash_coa else None,
            "active": True,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })
    await db.bank_accounts.insert_many(bank_docs)
    for d in bank_docs:
        d.pop("_id", None)

    PAYMENT_METHODS = [
        ("CASH", "Cash", "cash"),
        ("PETTY-CASH", "Petty Cash", "cash"),
        ("CONTRA", "Contra (Kontra Bon)", "transfer"),
        ("BCA-TRF", "BCA Transfer", "transfer"),
        ("MANDIRI-TRF", "Mandiri Transfer", "transfer"),
        ("BNI-TRF", "BNI Transfer", "transfer"),
        ("QRIS", "QRIS", "qris"),
        ("DEBIT-CARD", "Debit Card", "card"),
        ("CREDIT-CARD", "Credit Card", "card"),
        ("GOPAY", "GoPay", "qris"),
        ("OVO", "OVO", "qris"),
        ("DANA", "DANA", "qris"),
    ]
    pm_docs = []
    for code, name, ptype in PAYMENT_METHODS:
        pm_docs.append({
            "id": str(uuid.uuid4()),
            "code": code,
            "name": name,
            "type": ptype,
            "active": True,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })
    await db.payment_methods.insert_many(pm_docs)
    for d in pm_docs:
        d.pop("_id", None)
    log.info(f"[A.5] Inserted {len(bank_docs)} bank accounts + {len(pm_docs)} payment methods")

    log.info("=" * 70)
    log.info("PHASE A COMPLETE")
    log.info("=" * 70)


async def main():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    await phase_a_master(db)
    client.close()


if __name__ == "__main__":
    asyncio.run(main())
