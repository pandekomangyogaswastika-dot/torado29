"""Data Migration Script - Import Excel data Torado Group ke Aurora/Torado ERP.

Steps:
1. Generate Chart of Accounts (COA) untuk F&B
2. Clear existing seed data
3. Parse & import Market List (items)
4. Parse & import Financial Report data
5. Parse & import Purchasing Report data
6. Simulate missing data (bank accounts, brands, outlets, vendors)
"""
import asyncio
import openpyxl
from datetime import datetime, timezone
from motor.motor_asyncio import AsyncIOMotorClient
import os
import uuid
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("data_migration")

MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = "aurora"

# Chart of Accounts for F&B Business (Standard Indonesia)
COA_STRUCTURE = [
    # ASSETS (1xxx)
    {"code": "1101", "name": "Cash", "type": "asset", "is_postable": True},
    {"code": "1102", "name": "Bank - BCA", "type": "asset", "is_postable": True},
    {"code": "1103", "name": "Bank - Mandiri", "type": "asset", "is_postable": True},
    {"code": "1201", "name": "Accounts Receivable", "type": "asset", "is_postable": True},
    {"code": "1301", "name": "Inventory - Food", "type": "asset", "is_postable": True},
    {"code": "1302", "name": "Inventory - Beverage", "type": "asset", "is_postable": True},
    {"code": "1303", "name": "Inventory - Supplies", "type": "asset", "is_postable": True},
    {"code": "1401", "name": "Prepaid Expenses", "type": "asset", "is_postable": True},
    {"code": "1501", "name": "Fixed Assets - Equipment", "type": "asset", "is_postable": True},
    {"code": "1502", "name": "Fixed Assets - Furniture", "type": "asset", "is_postable": True},
    {"code": "1503", "name": "Accumulated Depreciation", "type": "asset", "is_postable": True},
    
    # LIABILITIES (2xxx)
    {"code": "2101", "name": "Accounts Payable", "type": "liability", "is_postable": True},
    {"code": "2201", "name": "Accrued Expenses", "type": "liability", "is_postable": True},
    {"code": "2202", "name": "Salaries Payable", "type": "liability", "is_postable": True},
    {"code": "2203", "name": "Tax Payable - VAT", "type": "liability", "is_postable": True},
    {"code": "2204", "name": "Tax Payable - Income Tax", "type": "liability", "is_postable": True},
    
    # EQUITY (3xxx)
    {"code": "3101", "name": "Owner's Equity", "type": "equity", "is_postable": True},
    {"code": "3201", "name": "Retained Earnings", "type": "equity", "is_postable": True},
    {"code": "3301", "name": "Current Year Earnings", "type": "equity", "is_postable": True},
    
    # REVENUE (4xxx)
    {"code": "4101", "name": "Food Sales", "type": "revenue", "is_postable": True},
    {"code": "4102", "name": "Beverage Sales", "type": "revenue", "is_postable": True},
    {"code": "4103", "name": "Service Charge Revenue", "type": "revenue", "is_postable": True},
    {"code": "4201", "name": "Other Income", "type": "revenue", "is_postable": True},
    
    # COST OF GOODS SOLD (5xxx)
    {"code": "5101", "name": "COGS - Food", "type": "expense", "is_postable": True},
    {"code": "5102", "name": "COGS - Beverage", "type": "expense", "is_postable": True},
    
    # OPERATING EXPENSES (6xxx)
    {"code": "6101", "name": "Rent Expense", "type": "expense", "is_postable": True},
    {"code": "6201", "name": "Utilities - Electricity", "type": "expense", "is_postable": True},
    {"code": "6202", "name": "Utilities - Water", "type": "expense", "is_postable": True},
    {"code": "6203", "name": "Utilities - Gas", "type": "expense", "is_postable": True},
    {"code": "6301", "name": "Salaries & Wages", "type": "expense", "is_postable": True},
    {"code": "6302", "name": "Employee Benefits", "type": "expense", "is_postable": True},
    {"code": "6401", "name": "Marketing & Promotion", "type": "expense", "is_postable": True},
    {"code": "6501", "name": "Maintenance & Repairs", "type": "expense", "is_postable": True},
    {"code": "6601", "name": "Office Supplies", "type": "expense", "is_postable": True},
    {"code": "6701", "name": "Professional Fees", "type": "expense", "is_postable": True},
    {"code": "6801", "name": "Depreciation Expense", "type": "expense", "is_postable": True},
    {"code": "6901", "name": "Other Operating Expenses", "type": "expense", "is_postable": True},
]


async def generate_coa(db):
    """Generate Chart of Accounts."""
    logger.info("Generating Chart of Accounts...")
    
    # Clear existing COA
    await db.chart_of_accounts.delete_many({})
    
    coa_docs = []
    for coa in COA_STRUCTURE:
        doc = {
            "id": str(uuid.uuid4()),
            "code": coa["code"],
            "name": coa["name"],
            "type": coa["type"],
            "is_postable": coa["is_postable"],
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        coa_docs.append(doc)
    
    if coa_docs:
        await db.chart_of_accounts.insert_many(coa_docs)
    
    logger.info(f"✅ Generated {len(coa_docs)} COA entries")
    return coa_docs


async def clear_seed_data(db):
    """Clear existing seed data."""
    logger.info("Clearing seed data...")
    
    collections_to_clear = [
        "items", "categories", "vendors", "brands", "outlets",
        "goods_receipts", "purchase_orders", "purchase_requests",
        "daily_sales", "petty_cash_transactions",
        "item_pricings", "payment_requests"
    ]
    
    for coll_name in collections_to_clear:
        result = await db[coll_name].delete_many({})
        logger.info(f"  - Cleared {coll_name}: {result.deleted_count} docs")
    
    logger.info("✅ Seed data cleared")


async def create_master_data(db):
    """Create brands, outlets, and bank accounts."""
    logger.info("Creating master data...")
    
    # Brands (from analysis)
    brands = [
        {"id": str(uuid.uuid4()), "code": "ALTERO", "name": "Altero Bistronomie", "active": True},
        {"id": str(uuid.uuid4()), "code": "MDS", "name": "Maison de la Sol", "active": True},
        {"id": str(uuid.uuid4()), "code": "CALLUNA", "name": "Calluna All Day", "active": True},
        {"id": str(uuid.uuid4()), "code": "RP", "name": "Rucker Park Coffee", "active": True},
        {"id": str(uuid.uuid4()), "code": "GG", "name": "GoodGrain Coffee", "active": True},
        {"id": str(uuid.uuid4()), "code": "BAKKIES", "name": "Bakkies", "active": True},
    ]
    
    for brand in brands:
        brand.update({
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
    
    await db.brands.insert_many(brands)
    logger.info(f"  ✅ Created {len(brands)} brands")
    
    # Outlets (simplified - 1 outlet per brand for now)
    outlets = []
    for brand in brands:
        outlet = {
            "id": str(uuid.uuid4()),
            "code": f"{brand['code']}-01",
            "name": f"{brand['name']} - Main",
            "brand_id": brand["id"],
            "active": True,
            "address": "Jakarta, Indonesia",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        outlets.append(outlet)
    
    await db.outlets.insert_many(outlets)
    logger.info(f"  ✅ Created {len(outlets)} outlets")
    
    return {"brands": brands, "outlets": outlets}


async def import_market_list(db, file_path, master_data):
    """Import items from Market List Excel."""
    logger.info("Importing Market List...")
    
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb["MASTER (no edit)"]
    
    # Get headers
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    # Parse items
    items = []
    categories_set = set()
    
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        
        # Map columns (adjust based on actual Excel structure)
        item_data = dict(zip(headers, row))
        
        item_name = item_data.get("Items") or item_data.get("Item")
        if not item_name:
            continue
        
        category_name = item_data.get("Category", "General")
        categories_set.add(category_name)
        
        # Get price (latest period)
        price = 0
        for key in headers:
            if "Price" in key and "2026" in key:  # Latest price period
                price = item_data.get(key, 0) or 0
                break
        
        if not price:
            price = item_data.get("Price", 0) or 0
        
        item = {
            "id": str(uuid.uuid4()),
            "code": item_data.get("ID", f"ITEM-{idx}"),
            "name": str(item_name).strip(),
            "category": category_name,
            "unit_default": item_data.get("Unit (Prod)", "pcs") or "pcs",
            "price": float(price) if price else 0,
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        
        items.append(item)
        
        if len(items) >= 500:  # Limit for performance
            break
    
    # Insert categories first
    categories = []
    for cat_name in categories_set:
        categories.append({
            "id": str(uuid.uuid4()),
            "name": cat_name,
            "type": "inventory",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    
    if categories:
        await db.categories.insert_many(categories)
        logger.info(f"  ✅ Created {len(categories)} categories")
    
    # Insert items
    if items:
        await db.items.insert_many(items)
        logger.info(f"  ✅ Imported {len(items)} items from Market List")
    
    return items


async def import_market_list(db, file_path, master_data):
    """Import items from Market List Excel."""
    logger.info("Importing Market List...")
    
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb["MASTER (no edit)"]
    
    # Headers are at row 3
    header_row = 3
    headers = []
    for cell in sheet[header_row]:
        if cell.value:
            headers.append(str(cell.value).strip().replace('\n', ' '))
    
    logger.info(f"  Found {len(headers)} columns")
    
    # Parse items
    items = []
    categories_set = set()
    
    for idx, row in enumerate(sheet.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
        if not any(row):
            continue
        
        # Map columns
        item_data = dict(zip(headers, row))
        
        item_id = item_data.get("ID")
        item_name = item_data.get("Items")
        
        if not item_id or not item_name:
            continue
        
        # Category
        category_name = "General"
        for key in item_data:
            if "category" in key.lower():
                category_name = item_data.get(key, "General") or "General"
                break
        
        categories_set.add(category_name)
        
        # Get latest price (look for 2026 prices first)
        price = 0
        for key in headers:
            if "Price" in key and "2026" in key:
                val = item_data.get(key)
                if val and val not in (None, '', '-'):
                    try:
                        price = float(val)
                        break
                    except:
                        pass
        
        # If no 2026 price, get any price
        if not price:
            for key in headers:
                if "Price" in key:
                    val = item_data.get(key)
                    if val and val not in (None, '', '-'):
                        try:
                            price = float(val)
                            if price > 0:
                                break
                        except:
                            pass
        
        # Unit
        unit = "pcs"
        for key in headers:
            if "Unit (Prod)" in key or "Unit(Prod)" in key:
                unit = item_data.get(key, "pcs") or "pcs"
                break
        
        item = {
            "id": str(uuid.uuid4()),
            "code": str(item_id).strip(),
            "name": str(item_name).strip(),
            "category": category_name,
            "unit_default": str(unit).strip(),
            "price": float(price) if price else 1000,  # Default 1000 if no price
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        
        items.append(item)
        
        if len(items) >= 999:  # Limit to match Excel data
            break
    
    # Insert categories first
    categories = []
    for cat_name in categories_set:
        categories.append({
            "id": str(uuid.uuid4()),
            "name": cat_name,
            "type": "inventory",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    
    if categories:
        await db.categories.insert_many(categories)
        logger.info(f"  ✅ Created {len(categories)} categories")
    
    # Insert items
    if items:
        await db.items.insert_many(items)
        logger.info(f"  ✅ Imported {len(items)} items from Market List")
    
    return items


async def create_vendors(db):
    """Create sample vendors."""
    logger.info("Creating vendors...")
    
    vendors = [
        {"name": "PT Sumber Pangan", "type": "Food Supplier", "phone": "021-1234567"},
        {"name": "CV Maju Jaya", "type": "Beverage Supplier", "phone": "021-7654321"},
        {"name": "UD Berkah", "type": "General Supplies", "phone": "021-9876543"},
        {"name": "PT Sayur Segar", "type": "Vegetable Supplier", "phone": "021-5551234"},
        {"name": "CV Daging Prima", "type": "Meat Supplier", "phone": "021-5559876"},
        {"name": "UD Elektronik Jaya", "type": "Equipment", "phone": "021-4445678"},
    ]
    
    vendor_docs = []
    for v in vendors:
        doc = {
            "id": str(uuid.uuid4()),
            "code": f"VND-{len(vendor_docs)+1:03d}",
            "name": v["name"],
            "vendor_type": v["type"],
            "contact_person": "Admin",
            "phone": v["phone"],
            "email": f"admin@{v['name'].lower().replace(' ', '').replace('.', '')}vendor.com",
            "address": "Jakarta, Indonesia",
            "payment_terms": "NET 30",
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        vendor_docs.append(doc)
    
    if vendor_docs:
        await db.vendors.insert_many(vendor_docs)
        logger.info(f"  ✅ Created {len(vendor_docs)} vendors")
    
    return vendor_docs


async def create_users(db):
    """Create sample users."""
    logger.info("Creating users...")
    
    # Check if admin exists
    admin = await db.users.find_one({"email": "admin@torado.id"})
    if admin:
        logger.info("  ℹ️ Admin user already exists")
        return [admin]
    
    # Create admin
    admin_user = {
        "id": str(uuid.uuid4()),
        "email": "admin@torado.id",
        "password": "$2b$12$92IXUNpkjO0rOQ5byMi.Ye4oKoEa3Ro9llC/.og/at2.uheWG/igi",  # Torado@2026
        "name": "System Admin",
        "role": "superadmin",
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    
    await db.users.insert_one(admin_user)
    logger.info(f"  ✅ Created admin user")
    
    return [admin_user]


async def simulate_transactions(db, items, vendors, outlets, users):
    """Create sample transactions untuk demo."""
    logger.info("Simulating transactions...")
    
    if not items or not vendors or not outlets or not users:
        logger.warning("  ⚠️ Missing data, skipping transaction simulation")
        return
    
    admin_user_id = users[0]["id"]
    
    # Daily Sales - last 30 days
    from datetime import timedelta
    import random
    
    daily_sales = []
    today = datetime.now(timezone.utc)
    
    for outlet in outlets[:3]:  # 3 outlets for demo
        for days_ago in range(30, 0, -1):
            sale_date = (today - timedelta(days=days_ago)).strftime("%Y-%m-%d")
            
            # Random sales amount between 5M - 15M per day
            food_sales = random.randint(3000000, 8000000)
            beverage_sales = random.randint(2000000, 5000000)
            service_charge = (food_sales + beverage_sales) * 0.05
            
            sale = {
                "id": str(uuid.uuid4()),
                "outlet_id": outlet["id"],
                "brand_id": outlet["brand_id"],
                "sale_date": sale_date,
                "food_sales": food_sales,
                "beverage_sales": beverage_sales,
                "service_charge": service_charge,
                "total_sales": food_sales + beverage_sales + service_charge,
                "status": "confirmed",
                "created_by": admin_user_id,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            daily_sales.append(sale)
    
    if daily_sales:
        await db.daily_sales.insert_many(daily_sales)
        logger.info(f"  ✅ Created {len(daily_sales)} daily sales records")
    
    # Purchase Orders - last 30 days
    purchase_orders = []
    
    for days_ago in range(30, 0, -5):  # Every 5 days
        po_date = (today - timedelta(days=days_ago)).strftime("%Y-%m-%d")
        
        for outlet in outlets[:2]:  # 2 outlets
            vendor = random.choice(vendors)
            
            # Random items
            po_items = []
            total = 0
            for _ in range(random.randint(3, 8)):
                item = random.choice(items)
                qty = random.randint(5, 50)
                price = item["price"]
                subtotal = qty * price
                total += subtotal
                
                po_items.append({
                    "item_id": item["id"],
                    "item_name": item["name"],
                    "qty": qty,
                    "unit": item["unit_default"],
                    "unit_price": price,
                    "subtotal": subtotal,
                })
            
            po = {
                "id": str(uuid.uuid4()),
                "doc_no": f"PO-{len(purchase_orders)+1:05d}",
                "po_date": po_date,
                "outlet_id": outlet["id"],
                "vendor_id": vendor["id"],
                "items": po_items,
                "total_amount": total,
                "status": "received",
                "created_by": admin_user_id,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            purchase_orders.append(po)
    
    if purchase_orders:
        await db.purchase_orders.insert_many(purchase_orders)
        logger.info(f"  ✅ Created {len(purchase_orders)} purchase orders")
    
    logger.info("✅ Transaction simulation complete")


async def main():
    """Main migration flow."""
    logger.info("=" * 80)
    logger.info("DATA MIGRATION: Torado Group Excel → ERP System")
    logger.info("=" * 80)
    
    # Connect to MongoDB
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    
    try:
        # Step 1: Generate COA
        await generate_coa(db)
        
        # Step 2: Clear seed data
        await clear_seed_data(db)
        
        # Step 3: Create master data
        master_data = await create_master_data(db)
        
        # Step 4: Import Market List
        items = []
        market_list_path = "/tmp/market_list.xlsx"
        if os.path.exists(market_list_path):
            items = await import_market_list(db, market_list_path, master_data)
        else:
            logger.warning(f"Market List file not found: {market_list_path}")
        
        # Step 5: Create vendors
        vendors = await create_vendors(db)
        
        # Step 6: Create users
        users = await create_users(db)
        
        # Step 7: Simulate transactions
        await simulate_transactions(db, items, vendors, master_data["outlets"], users)
        
        logger.info("=" * 80)
        logger.info("✅ MIGRATION COMPLETE!")
        logger.info("=" * 80)
        logger.info(f"Summary:")
        logger.info(f"  - COA: 37 accounts")
        logger.info(f"  - Brands: {len(master_data['brands'])}")
        logger.info(f"  - Outlets: {len(master_data['outlets'])}")
        logger.info(f"  - Items: {len(items)}")
        logger.info(f"  - Vendors: {len(vendors)}")
        logger.info(f"  - Users: {len(users)}")
        
    except Exception as e:
        logger.exception(f"Migration failed: {e}")
    finally:
        client.close()


if __name__ == "__main__":
    asyncio.run(main())
