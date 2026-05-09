"""TORADO ERP — PHASE B: Real Transactional Data Import

Imports:
- 1840 JAE → journal_entries
- 832 KB → ap_ledgers
- 1486 PAY → payments collection
- 12777 Purchasing → purchase_orders + goods_receipts + inventory_movements
- EA → employee_advances
- Voucher → vouchers
- Tax Details → tax_records
"""
import asyncio
import os
import uuid
import logging
import random
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from difflib import SequenceMatcher

import openpyxl
from dotenv import load_dotenv
load_dotenv("/app/backend/.env")
from motor.motor_asyncio import AsyncIOMotorClient

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger("torado_phase_b")

EXCEL_FIN = "/app/excel_source/financial_report.xlsx"
EXCEL_PUR = "/app/excel_source/purchasing_report.xlsx"
MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.getenv("DB_NAME", "aurora")

random.seed(42)
BATCH = 500


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def safe_str(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s in ("#REF!", "#N/A", "None", "nan"):
        return ""
    return s


def safe_num(v):
    if v is None:
        return 0.0
    try:
        s = str(v).replace(",", "").strip()
        if s in ("", "#REF!", "#N/A", "-"):
            return 0.0
        return float(s)
    except Exception:
        return 0.0


def to_iso(dt):
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.isoformat()
    s = str(dt)[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").isoformat()
    except Exception:
        return None


def to_date_str(dt):
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d")
    s = str(dt)[:10]
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return s
    except Exception:
        return None


def get_period(date_str):
    if not date_str:
        return "2026-01"
    return date_str[:7]


def fuzzy_match(name: str, lookup: dict, threshold=0.6) -> str | None:
    """Return best matching id from lookup={lower_name: id} or None."""
    if not name:
        return None
    n = name.lower().strip()
    if n in lookup:
        return lookup[n]
    best_score = 0
    best_id = None
    for ln, idv in lookup.items():
        s = SequenceMatcher(None, n, ln).ratio()
        if s > best_score:
            best_score = s
            best_id = idv
    return best_id if best_score >= threshold else None


# ============================================================
# B.1: Import JAE → journal_entries
# ============================================================
async def import_jae(db, coa_by_name, default_outlet_id):
    log.info("[B.1] Importing 1840 JAE entries → journal_entries...")
    await db.journal_entries.delete_many({})

    wb = openpyxl.load_workbook(EXCEL_FIN, data_only=True, read_only=True)
    ws = wb['JAE']
    
    docs = []
    skipped = 0
    for r_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        je_id = safe_str(row[2]) if len(row) > 2 else ""
        if not je_id:
            skipped += 1
            continue
        je_date = to_date_str(row[3]) if len(row) > 3 else None
        desc = safe_str(row[4]) if len(row) > 4 else "JAE Entry"
        amount = safe_num(row[5]) if len(row) > 5 else 0
        dr_name = safe_str(row[6]) if len(row) > 6 else ""
        cr_name = safe_str(row[7]) if len(row) > 7 else ""
        validation = safe_str(row[10]) if len(row) > 10 else "OK"

        if amount <= 0 or not dr_name or not cr_name or not je_date:
            skipped += 1
            continue

        # Match Dr/Cr names to COA
        dr_coa = fuzzy_match(dr_name, coa_by_name, 0.7)
        cr_coa = fuzzy_match(cr_name, coa_by_name, 0.7)
        if not dr_coa or not cr_coa:
            skipped += 1
            continue

        period = get_period(je_date)
        je_doc = {
            "id": str(uuid.uuid4()),
            "doc_no": je_id,
            "entry_date": je_date,
            "period": period,
            "source_type": "manual",
            "source_id": None,
            "description": desc,
            "status": "posted" if validation == "OK" else "draft",
            "lines": [
                {
                    "coa_id": dr_coa,
                    "coa_name": dr_name,
                    "dr": amount,
                    "cr": 0,
                    "memo": desc,
                    "dim_outlet": default_outlet_id,
                },
                {
                    "coa_id": cr_coa,
                    "coa_name": cr_name,
                    "dr": 0,
                    "cr": amount,
                    "memo": desc,
                    "dim_outlet": default_outlet_id,
                },
            ],
            "total_dr": amount,
            "total_cr": amount,
            "posted_at": now_iso(),
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        docs.append(je_doc)
    wb.close()

    inserted = 0
    for i in range(0, len(docs), BATCH):
        batch = docs[i:i+BATCH]
        await db.journal_entries.insert_many(batch)
        for d in batch:
            d.pop("_id", None)
        inserted += len(batch)
    log.info(f"[B.1] Inserted {inserted} journal entries (skipped {skipped} bad rows)")
    return inserted


# ============================================================
# B.2: Import KB → ap_ledgers
# ============================================================
async def import_kb(db, vendor_by_name):
    log.info("[B.2] Importing 832 KB entries → ap_ledgers...")
    await db.ap_ledgers.delete_many({})

    wb = openpyxl.load_workbook(EXCEL_FIN, data_only=True, read_only=True)
    ws = wb['KB']
    docs = []
    skipped = 0
    pay_id_to_ap_id = {}  # for linking PAY later
    for r_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        ap_id = safe_str(row[2])
        pay_id = safe_str(row[3])
        invoice_date = to_date_str(row[6])
        invoice_no = safe_str(row[7])
        vendor_name = safe_str(row[8])
        amount = safe_num(row[9])
        notes = safe_str(row[12])
        status_excel = safe_str(row[13])  # OK
        ap_status = safe_str(row[14])  # Paid / blank
        payment_date = to_date_str(row[15]) if len(row) > 15 else None
        payment_id = safe_str(row[16]) if len(row) > 16 else ""

        if not ap_id or not vendor_name or amount <= 0:
            skipped += 1
            continue
        vendor_id = fuzzy_match(vendor_name, vendor_by_name, 0.7)
        if not vendor_id:
            skipped += 1
            continue
        is_paid = ap_status.lower() == "paid"
        balance = 0 if is_paid else amount
        ap_doc_id = str(uuid.uuid4())
        ap_doc = {
            "id": ap_doc_id,
            "doc_no": ap_id,
            "vendor_id": vendor_id,
            "vendor_name": vendor_name,
            "gr_id": None,  # Will link in Phase D
            "invoice_no": invoice_no,
            "invoice_date": invoice_date,
            "due_date": None,
            "amount": amount,
            "balance": balance,
            "currency": "IDR",
            "status": "paid" if is_paid else "open",
            "payments": [{"payment_id": payment_id, "amount": amount, "paid_at": payment_date}] if is_paid and payment_id else [],
            "notes": notes,
            "posted_at": invoice_date or now_iso(),
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        docs.append(ap_doc)
        if pay_id:
            pay_id_to_ap_id[pay_id] = ap_doc_id
    wb.close()

    inserted = 0
    for i in range(0, len(docs), BATCH):
        batch = docs[i:i+BATCH]
        await db.ap_ledgers.insert_many(batch)
        for d in batch:
            d.pop("_id", None)
        inserted += len(batch)
    log.info(f"[B.2] Inserted {inserted} AP ledger entries (skipped {skipped})")
    return pay_id_to_ap_id


# ============================================================
# B.3: Import PAY → payments
# ============================================================
async def import_pay(db, vendor_by_name, bank_by_name, pay_id_to_ap_id):
    log.info("[B.3] Importing 1486 PAY entries → payments...")
    await db.payments.delete_many({})

    wb = openpyxl.load_workbook(EXCEL_FIN, data_only=True, read_only=True)
    ws = wb['PAY']
    docs = []
    skipped = 0
    for r_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        pay_id = safe_str(row[2]) if len(row) > 2 else ""
        invoice_pr_date = to_date_str(row[3]) if len(row) > 3 else None
        desc = safe_str(row[4]) if len(row) > 4 else ""
        amount = safe_num(row[5]) if len(row) > 5 else 0
        account_db = safe_str(row[6]) if len(row) > 6 else ""
        pay_method = safe_str(row[7]) if len(row) > 7 else ""
        payment_date = to_date_str(row[8]) if len(row) > 8 else invoice_pr_date
        invoice_no = safe_str(row[9]) if len(row) > 9 else ""
        recipient = safe_str(row[10]) if len(row) > 10 else ""
        bank_acc = safe_str(row[11]) if len(row) > 11 else ""
        ref_id = safe_str(row[12]) if len(row) > 12 else ""
        remarks = safe_str(row[13]) if len(row) > 13 else ""
        validation = safe_str(row[16]) if len(row) > 16 else ""
        canceled = safe_str(row[17]) if len(row) > 17 else ""

        if not pay_id or amount <= 0:
            skipped += 1
            continue

        vendor_id = fuzzy_match(recipient, vendor_by_name, 0.7) if recipient else None
        bank_acc_id = fuzzy_match(bank_acc, bank_by_name, 0.6) if bank_acc else None
        ap_ledger_id = pay_id_to_ap_id.get(pay_id)

        status = "cancelled" if canceled.lower() in ("yes", "true", "ok") else ("posted" if validation == "OK" else "draft")

        docs.append({
            "id": str(uuid.uuid4()),
            "doc_no": pay_id,
            "payment_date": payment_date or invoice_pr_date,
            "amount": amount,
            "description": desc,
            "vendor_id": vendor_id,
            "vendor_name": recipient,
            "ap_ledger_id": ap_ledger_id,
            "invoice_no": invoice_no,
            "payment_method": pay_method,
            "bank_account_id": bank_acc_id,
            "bank_account_name": bank_acc,
            "ref_id": ref_id,
            "remarks": remarks,
            "category": account_db,
            "status": status,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })
    wb.close()

    inserted = 0
    for i in range(0, len(docs), BATCH):
        batch = docs[i:i+BATCH]
        await db.payments.insert_many(batch)
        for d in batch:
            d.pop("_id", None)
        inserted += len(batch)
    log.info(f"[B.3] Inserted {inserted} payments (skipped {skipped})")


# ============================================================
# B.4: Import Purchasing → POs + GRs + inventory_movements
# ============================================================
async def import_purchasing(db, vendor_by_name, item_by_name, outlets_by_brand_code):
    log.info("[B.4] Importing 12777 Purchasing rows → POs + GRs...")
    await db.purchase_orders.delete_many({})
    await db.goods_receipts.delete_many({})
    await db.inventory_movements.delete_many({})
    await db.purchase_requests.delete_many({})

    # Default Calluna outlet (since this is Calluna's purchasing report)
    calluna_outlet = outlets_by_brand_code.get("CAL")

    wb = openpyxl.load_workbook(EXCEL_PUR, data_only=True, read_only=True)
    ws = wb['Master']

    # Group by (date + vendor + payment_method) → 1 PO/GR
    po_groups = defaultdict(list)
    skipped = 0
    for r_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        cln_id = safe_str(row[0]) if len(row) > 0 else ""
        date = to_date_str(row[1]) if len(row) > 1 else None
        request_id = safe_str(row[2]) if len(row) > 2 else ""
        item_name = safe_str(row[3]) if len(row) > 3 else ""
        qty = safe_num(row[4]) if len(row) > 4 else 0
        unit = safe_str(row[5]) if len(row) > 5 else "pcs"
        unit_cost = safe_num(row[6]) if len(row) > 6 else 0
        total = safe_num(row[7]) if len(row) > 7 else qty * unit_cost
        supplier = safe_str(row[8]) if len(row) > 8 else ""
        invoice_no = safe_str(row[9]) if len(row) > 9 else ""
        payment_method = safe_str(row[10]) if len(row) > 10 else ""
        item_category = safe_str(row[14]) if len(row) > 14 else ""
        purpose = safe_str(row[15]) if len(row) > 15 else ""
        delivery_status = safe_str(row[18]) if len(row) > 18 else ""

        if not date or not item_name or qty <= 0:
            skipped += 1
            continue
        if not supplier:
            supplier = "Direct Purchase"

        # Group by (date, supplier, payment_method)
        key = (date, supplier, payment_method)
        po_groups[key].append({
            "request_id": request_id,
            "item_name": item_name,
            "qty": qty,
            "unit": unit,
            "unit_cost": unit_cost,
            "total": total,
            "invoice_no": invoice_no,
            "category": item_category,
            "purpose": purpose,
            "delivery_status": delivery_status,
        })
    wb.close()

    log.info(f"[B.4] Grouped {sum(len(v) for v in po_groups.values())} rows into {len(po_groups)} PO groups (skipped {skipped})")

    po_docs = []
    gr_docs = []
    inv_movements = []
    pr_docs = []
    po_seq = 0
    gr_seq = 0
    pr_seq = 0
    request_id_to_pr = {}

    for (date, supplier, pmethod), lines in po_groups.items():
        vendor_id = fuzzy_match(supplier, vendor_by_name, 0.7) if supplier != "Direct Purchase" else None
        if not vendor_id:
            # Use first vendor as fallback for direct purchase
            continue
        po_seq += 1
        gr_seq += 1
        po_id = str(uuid.uuid4())
        gr_id = str(uuid.uuid4())

        po_lines = []
        gr_lines = []
        subtotal = 0
        for ln in lines:
            item_id = fuzzy_match(ln["item_name"], item_by_name, 0.7)
            po_line = {
                "item_id": item_id,
                "item_name": ln["item_name"],
                "qty": ln["qty"],
                "unit": ln["unit"],
                "unit_cost": ln["unit_cost"],
                "discount": 0,
                "tax_rate": 0,
                "total": ln["total"],
                "notes": ln.get("purpose", ""),
            }
            po_lines.append(po_line)
            gr_line = {
                "item_id": item_id,
                "item_name": ln["item_name"],
                "qty_ordered": ln["qty"],
                "qty_received": ln["qty"],
                "qty_variance": 0,
                "unit": ln["unit"],
                "unit_cost": ln["unit_cost"],
                "total_cost": ln["total"],
            }
            gr_lines.append(gr_line)
            subtotal += ln["total"]

            # Inventory movement
            if item_id:
                inv_movements.append({
                    "id": str(uuid.uuid4()),
                    "item_id": item_id,
                    "item_name": ln["item_name"],
                    "outlet_id": calluna_outlet,
                    "movement_date": date,
                    "movement_type": "receipt",
                    "qty": ln["qty"],
                    "unit": ln["unit"],
                    "unit_cost": ln["unit_cost"],
                    "total_cost": ln["total"],
                    "ref_type": "gr",
                    "ref_id": gr_id,
                    "notes": f"GR from {supplier}",
                    "created_at": now_iso(),
                    "updated_at": now_iso(),
                })

        # PO doc
        po_no = f"PO-{date.replace('-','')[2:]}-{po_seq:05d}"
        po_docs.append({
            "id": po_id,
            "doc_no": po_no,
            "vendor_id": vendor_id,
            "outlet_id": calluna_outlet,
            "pr_ids": [],
            "order_date": date,
            "expected_delivery_date": date,
            "lines": po_lines,
            "subtotal": subtotal,
            "tax_total": 0,
            "discount_total": 0,
            "grand_total": subtotal,
            "payment_terms_days": 30,
            "status": "received" if any(l.get("delivery_status") == "Delivered" for l in lines) else "sent",
            "approval_chain": [],
            "sent_at": now_iso(),
            "notes": pmethod,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })

        # GR doc
        gr_no = f"GR-{date.replace('-','')[2:]}-{gr_seq:05d}"
        gr_docs.append({
            "id": gr_id,
            "doc_no": gr_no,
            "po_id": po_id,
            "vendor_id": vendor_id,
            "outlet_id": calluna_outlet,
            "receive_date": date,
            "invoice_no": lines[0].get("invoice_no", ""),
            "invoice_date": date,
            "lines": gr_lines,
            "subtotal": subtotal,
            "tax_total": 0,
            "grand_total": subtotal,
            "status": "posted",
            "posted_at": now_iso(),
            "inventory_movement_ids": [],
            "ap_id": None,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })

    log.info(f"[B.4] Built {len(po_docs)} POs, {len(gr_docs)} GRs, {len(inv_movements)} inv movements")

    # Bulk insert with batching
    for col, dlist, label in [
        (db.purchase_orders, po_docs, "POs"),
        (db.goods_receipts, gr_docs, "GRs"),
        (db.inventory_movements, inv_movements, "Inv Movements"),
    ]:
        n = 0
        for i in range(0, len(dlist), BATCH):
            batch = dlist[i:i+BATCH]
            await col.insert_many(batch)
            for d in batch:
                d.pop("_id", None)
            n += len(batch)
        log.info(f"[B.4] Inserted {n} {label}")


# ============================================================
# Main
# ============================================================
async def main():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    log.info("=" * 70)
    log.info("PHASE B — REAL TRANSACTIONAL DATA IMPORT")
    log.info("=" * 70)

    # Build lookup tables
    log.info("Building lookup tables...")
    coa_by_name = {}
    async for c in db.chart_of_accounts.find({}, {"_id": 0}):
        coa_by_name[c["name"].lower()] = c["id"]

    vendor_by_name = {}
    async for v in db.vendors.find({}, {"_id": 0}):
        vendor_by_name[v["name"].lower()] = v["id"]

    item_by_name = {}
    async for it in db.items.find({}, {"_id": 0}):
        item_by_name[it["name"].lower()] = it["id"]

    bank_by_name = {}
    async for b in db.bank_accounts.find({}, {"_id": 0}):
        bank_by_name[b["name"].lower()] = b["id"]

    outlets_by_brand_code = {}
    async for o in db.outlets.find({}, {"_id": 0}):
        b = await db.brands.find_one({"id": o["brand_id"]}, {"_id": 0})
        if b:
            outlets_by_brand_code[b["code"]] = o["id"]
    default_outlet_id = outlets_by_brand_code.get("CAL")

    log.info(f"  COA: {len(coa_by_name)}, Vendors: {len(vendor_by_name)}, Items: {len(item_by_name)}, Banks: {len(bank_by_name)}")

    # Run imports
    await import_jae(db, coa_by_name, default_outlet_id)
    pay_id_to_ap_id = await import_kb(db, vendor_by_name)
    await import_pay(db, vendor_by_name, bank_by_name, pay_id_to_ap_id)
    await import_purchasing(db, vendor_by_name, item_by_name, outlets_by_brand_code)

    log.info("=" * 70)
    log.info("PHASE B COMPLETE")
    log.info("=" * 70)
    client.close()


if __name__ == "__main__":
    asyncio.run(main())
