"""TORADO ERP — PHASE C: Smart Simulation Layer

Simulates missing transactional data so all dashboards have realistic content.

Simulates:
- Daily Sales: ~2000 records (365 days × 6 outlets, with realistic patterns)
- Petty Cash: ~300 transactions
- Customers: 80 realistic Indonesian customers
- Loyalty Transactions: ~400 records
- Cash Balance Snapshots: daily per bank account
- Anomaly Events: ~15 sample events
- Notifications: ~80 sample notifications
- Forecasts: pre-computed monthly projections
"""
import asyncio
import os
import uuid
import logging
import random
from datetime import datetime, timezone, timedelta
from collections import defaultdict

from dotenv import load_dotenv
load_dotenv("/app/backend/.env")
from motor.motor_asyncio import AsyncIOMotorClient

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger("torado_phase_c")

MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.getenv("DB_NAME", "aurora")

random.seed(42)
BATCH = 500


def now_iso():
    return datetime.now(timezone.utc).isoformat()


# Date range: full year 2026 + last 4 months 2025
START_DATE = datetime(2025, 9, 1)
END_DATE = datetime(2026, 5, 8)


# ============================================================
# C.1: Simulate Daily Sales
# ============================================================
async def simulate_daily_sales(db):
    log.info("[C.1] Simulating Daily Sales for full year × 6 outlets...")
    await db.daily_sales.delete_many({})

    outlets = []
    async for o in db.outlets.find({}, {"_id": 0}):
        outlets.append(o)
    brands_by_id = {}
    async for b in db.brands.find({}, {"_id": 0}):
        brands_by_id[b["id"]] = b

    payment_methods = []
    async for pm in db.payment_methods.find({"type": {"$in": ["cash", "transfer", "qris", "card"]}}, {"_id": 0}):
        payment_methods.append(pm)

    # Brand-specific baseline daily sales (in IDR) — DB codes: ALT, DLS, CAL, RKP, BKK
    BRAND_BASELINE = {
        "ALT": 8_500_000,    # Altero - premium bistro
        "DLS": 6_200_000,    # De La Sol - mid-tier resto
        "CAL": 12_000_000,   # Calluna - busy all-day
        "RKP": 4_500_000,    # Rucker Park - coffee
        "BKK": 5_500_000,    # Bakkies - casual
    }

    docs = []
    days = (END_DATE - START_DATE).days
    for d in range(days):
        date = START_DATE + timedelta(days=d)
        date_str = date.strftime("%Y-%m-%d")
        period = date_str[:7]
        is_weekend = date.weekday() >= 5
        is_payday = date.day in (1, 15, 25)
        for o in outlets:
            brand = brands_by_id.get(o["brand_id"])
            if not brand:
                continue
            baseline = BRAND_BASELINE.get(brand["code"], 5_000_000)
            # Weekly pattern + randomness
            weekend_mult = 1.4 if is_weekend else 1.0
            payday_mult = 1.2 if is_payday else 1.0
            random_mult = random.uniform(0.7, 1.3)
            # Seasonal - higher Oct-Dec, lower Jan-Feb
            month = date.month
            season = {1: 0.85, 2: 0.85, 3: 0.95, 4: 1.0, 5: 1.05, 6: 1.0,
                      7: 1.05, 8: 1.05, 9: 1.0, 10: 1.1, 11: 1.15, 12: 1.25}.get(month, 1.0)
            net_sales = round(baseline * weekend_mult * payday_mult * random_mult * season)
            # Skip ~3% randomly (closed days, holidays)
            if random.random() < 0.03:
                continue
            transaction_count = max(20, int(net_sales / random.randint(80_000, 180_000)))
            
            food_split = random.uniform(0.55, 0.70)
            beverage_split = random.uniform(0.20, 0.35)
            other_split = 1 - food_split - beverage_split

            food_sales = round(net_sales * food_split)
            bev_sales = round(net_sales * beverage_split)
            other_sales = net_sales - food_sales - bev_sales

            # Channels split
            dine_in_pct = random.uniform(0.50, 0.75) if not is_weekend else random.uniform(0.55, 0.80)
            takeaway_pct = random.uniform(0.10, 0.20)
            delivery_pct = 1 - dine_in_pct - takeaway_pct
            dine_in_sales = round(net_sales * dine_in_pct)
            takeaway_sales = round(net_sales * takeaway_pct)
            delivery_sales = net_sales - dine_in_sales - takeaway_sales

            channels = [
                {"channel": "dine_in", "gross": dine_in_sales, "discount": 0, "net": dine_in_sales},
                {"channel": "takeaway", "gross": takeaway_sales, "discount": 0, "net": takeaway_sales},
                {"channel": "online", "gross": delivery_sales, "discount": 0, "net": delivery_sales},
            ]
            revenue_buckets = [
                {"bucket": "food", "amount": food_sales},
                {"bucket": "beverage", "amount": bev_sales},
                {"bucket": "other", "amount": other_sales},
            ]
            
            # Payment breakdown
            cash_pct = random.uniform(0.20, 0.40)
            transfer_pct = random.uniform(0.10, 0.25)
            qris_pct = random.uniform(0.15, 0.30)
            card_pct = 1 - cash_pct - transfer_pct - qris_pct
            cash_amt = round(net_sales * cash_pct)
            transfer_amt = round(net_sales * transfer_pct)
            qris_amt = round(net_sales * qris_pct)
            card_amt = net_sales - cash_amt - transfer_amt - qris_amt

            cash_pm = next((p for p in payment_methods if p["code"] == "CASH"), None)
            qris_pm = next((p for p in payment_methods if p["code"] == "QRIS"), None)
            card_pm = next((p for p in payment_methods if p["code"] == "DEBIT-CARD"), None)
            transfer_pm = next((p for p in payment_methods if p["code"] == "BCA-TRF"), None)

            payment_breakdown = []
            if cash_pm: payment_breakdown.append({"payment_method_id": cash_pm["id"], "payment_method_name": cash_pm["name"], "amount": cash_amt})
            if transfer_pm: payment_breakdown.append({"payment_method_id": transfer_pm["id"], "payment_method_name": transfer_pm["name"], "amount": transfer_amt})
            if qris_pm: payment_breakdown.append({"payment_method_id": qris_pm["id"], "payment_method_name": qris_pm["name"], "amount": qris_amt})
            if card_pm: payment_breakdown.append({"payment_method_id": card_pm["id"], "payment_method_name": card_pm["name"], "amount": card_amt})

            service_charge = round(net_sales * 0.05)
            tax_amount = round((net_sales + service_charge) * 0.10)
            grand_total = net_sales + service_charge + tax_amount

            # Status: validated for older, mix for recent
            age_days = (END_DATE - date).days
            if age_days > 30:
                status = "validated"
            elif age_days > 7:
                status = random.choices(["validated", "submitted"], weights=[0.85, 0.15])[0]
            else:
                status = random.choices(["validated", "submitted", "draft"], weights=[0.6, 0.3, 0.1])[0]

            docs.append({
                "id": str(uuid.uuid4()),
                "outlet_id": o["id"],
                "brand_id": o["brand_id"],
                "sales_date": date_str,
                "status": status,
                "schema_version": 1,
                "channels": channels,
                "payment_breakdown": payment_breakdown,
                "revenue_buckets": revenue_buckets,
                "service_charge": service_charge,
                "tax_amount": tax_amount,
                "grand_total": grand_total,
                "transaction_count": transaction_count,
                "submitted_at": now_iso() if status != "draft" else None,
                "validated_at": now_iso() if status == "validated" else None,
                "created_at": now_iso(),
                "updated_at": now_iso(),
            })

    inserted = 0
    for i in range(0, len(docs), BATCH):
        batch = docs[i:i+BATCH]
        await db.daily_sales.insert_many(batch)
        for d in batch:
            d.pop("_id", None)
        inserted += len(batch)
    log.info(f"[C.1] Inserted {inserted} daily sales records")


# ============================================================
# C.2: Simulate Petty Cash
# ============================================================
async def simulate_petty_cash(db):
    log.info("[C.2] Simulating Petty Cash transactions...")
    await db.petty_cash.delete_many({})

    outlets = []
    async for o in db.outlets.find({}, {"_id": 0}):
        outlets.append(o)
    
    expense_descs = [
        ("Beli bumbu dapur (cabe, tomat, bawang)", "Vegetable Product"),
        ("Beli galon air minum", "Beverage Product"),
        ("Beli tisu, sabun cuci piring", "Cleaning Supplies"),
        ("Pembayaran kurir ekspedisi", "Operational"),
        ("Beli kantong sampah, sapu", "Cleaning Supplies"),
        ("Beli alat tulis kantor", "Office Supplies"),
        ("Pembayaran air galon", "Beverage Product"),
        ("Beli bahan dapur urgent", "Food Materials"),
        ("Top up token listrik", "Utilities"),
        ("Service AC bulanan", "Maintenance"),
        ("Beli gas LPG 12kg", "Utilities"),
        ("Pembayaran parkir tamu VIP", "Operational"),
        ("Beli es batu kemasan", "Beverage Product"),
        ("Reimbursement transport karyawan", "Transport"),
        ("Beli tinta printer", "Office Supplies"),
    ]
    docs = []
    days = (END_DATE - START_DATE).days
    seq = 0
    for d in range(0, days, 2):  # every other day
        date = START_DATE + timedelta(days=d)
        date_str = date.strftime("%Y-%m-%d")
        for o in outlets:
            # 0-3 PC txns per outlet per day
            n = random.choices([0, 1, 2, 3], weights=[0.3, 0.4, 0.2, 0.1])[0]
            for _ in range(n):
                desc, category = random.choice(expense_descs)
                amount = random.choice([25_000, 50_000, 75_000, 100_000, 125_000, 150_000, 200_000, 250_000, 300_000, 500_000, 750_000])
                seq += 1
                docs.append({
                    "id": str(uuid.uuid4()),
                    "doc_no": f"PC-{date.strftime('%y%m%d')}-{seq:04d}",
                    "outlet_id": o["id"],
                    "txn_date": date_str,
                    "type": "purchase",
                    "amount": amount,
                    "description": desc,
                    "vendor_text": random.choice(["Toko Makmur", "Indomaret", "Alfamart", "Pasar Tradisional", "Ace Hardware", "Toko Sentosa"]),
                    "category_id": None,
                    "status": "posted",
                    "balance_after": 0,
                    "created_at": now_iso(),
                    "updated_at": now_iso(),
                })

    inserted = 0
    for i in range(0, len(docs), BATCH):
        batch = docs[i:i+BATCH]
        await db.petty_cash.insert_many(batch)
        for d in batch:
            d.pop("_id", None)
        inserted += len(batch)
    log.info(f"[C.2] Inserted {inserted} petty cash transactions")


# ============================================================
# C.3: Simulate Customers
# ============================================================
async def simulate_customers(db):
    log.info("[C.3] Simulating customer master + loyalty data...")
    await db.customers.delete_many({})
    await db.loyalty_transactions.delete_many({})

    FIRST_NAMES = ["Adi", "Budi", "Citra", "Dian", "Eko", "Fitri", "Gina", "Hadi", "Indah",
                   "Joko", "Kartika", "Linda", "Maya", "Nia", "Oka", "Putri", "Rina", "Sari",
                   "Tina", "Umar", "Vera", "Wati", "Yani", "Zaki", "Ahmad", "Bambang"]
    LAST_NAMES = ["Wijaya", "Susanto", "Pratama", "Kusuma", "Lestari", "Hidayat", "Suryani",
                  "Anggraini", "Maulana", "Rahmawati", "Setiawan", "Putri", "Mahendra", "Sinaga"]
    TIERS = ["bronze", "silver", "gold", "platinum"]
    TIER_WEIGHTS = [0.50, 0.30, 0.15, 0.05]

    docs = []
    for i in range(80):
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        full_name = f"{first} {last}"
        tier = random.choices(TIERS, TIER_WEIGHTS)[0]
        points_balance = {"bronze": random.randint(0, 500),
                          "silver": random.randint(500, 2000),
                          "gold": random.randint(2000, 8000),
                          "platinum": random.randint(8000, 25000)}[tier]
        total_spent = points_balance * 1000
        join_date = (START_DATE + timedelta(days=random.randint(0, 200))).strftime("%Y-%m-%d")
        docs.append({
            "id": str(uuid.uuid4()),
            "code": f"CUST-{i+1:05d}",
            "full_name": full_name,
            "email": f"{first.lower()}.{last.lower()}{random.randint(1, 999)}@gmail.com",
            "phone": f"08{random.randint(10, 99)}-{random.randint(1000, 9999)}-{random.randint(1000, 9999)}",
            "tier": tier,
            "points_balance": points_balance,
            "total_spent": total_spent,
            "visits_count": random.randint(1, 50),
            "last_visit": (END_DATE - timedelta(days=random.randint(0, 60))).strftime("%Y-%m-%d"),
            "join_date": join_date,
            "address": f"Jl. {random.choice(['Mawar', 'Melati', 'Kenanga', 'Cempaka', 'Anggrek'])} No. {random.randint(1, 200)}, Jakarta",
            "city": random.choice(["Jakarta", "Bandung", "Surabaya", "Bali", "Yogyakarta"]),
            "active": True,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })
    if docs:
        await db.customers.insert_many(docs)
        for d in docs:
            d.pop("_id", None)
    log.info(f"[C.3] Inserted {len(docs)} customers")

    # Loyalty transactions
    customers = docs
    outlets = []
    async for o in db.outlets.find({}, {"_id": 0}):
        outlets.append(o)
    
    loyalty_txn = []
    for c in customers:
        n_txn = random.randint(2, 12)
        for _ in range(n_txn):
            txn_date = (datetime.strptime(c["join_date"], "%Y-%m-%d") + timedelta(days=random.randint(1, 200))).strftime("%Y-%m-%d")
            if txn_date > END_DATE.strftime("%Y-%m-%d"):
                continue
            outlet = random.choice(outlets)
            spend = random.choice([85_000, 120_000, 175_000, 250_000, 320_000, 450_000, 580_000])
            points_earned = int(spend / 1000)
            txn_type = random.choices(["earn", "redeem"], weights=[0.85, 0.15])[0]
            loyalty_txn.append({
                "id": str(uuid.uuid4()),
                "customer_id": c["id"],
                "customer_name": c["full_name"],
                "outlet_id": outlet["id"],
                "txn_date": txn_date,
                "type": txn_type,
                "spend_amount": spend if txn_type == "earn" else 0,
                "points_change": points_earned if txn_type == "earn" else -random.randint(100, points_earned * 2),
                "description": f"{'Earn' if txn_type == 'earn' else 'Redeem'} points - {outlet['name']}",
                "created_at": now_iso(),
                "updated_at": now_iso(),
            })
    if loyalty_txn:
        for i in range(0, len(loyalty_txn), BATCH):
            batch = loyalty_txn[i:i+BATCH]
            await db.loyalty_transactions.insert_many(batch)
            for d in batch:
                d.pop("_id", None)
    log.info(f"[C.3] Inserted {len(loyalty_txn)} loyalty transactions")


# ============================================================
# C.4: Cash Balance Snapshots
# ============================================================
async def simulate_cash_snapshots(db):
    log.info("[C.4] Simulating cash balance snapshots...")
    await db.cash_balance_snapshots.delete_many({})

    bank_accounts = []
    async for b in db.bank_accounts.find({}, {"_id": 0}):
        bank_accounts.append(b)

    docs = []
    days = (END_DATE - START_DATE).days
    # Initial balances per account
    initial_balances = {}
    for b in bank_accounts:
        initial_balances[b["id"]] = random.uniform(50_000_000, 500_000_000)

    for d in range(0, days, 7):  # weekly snapshots
        date = START_DATE + timedelta(days=d)
        date_str = date.strftime("%Y-%m-%d")
        for b in bank_accounts:
            balance = initial_balances[b["id"]] + random.uniform(-50_000_000, 80_000_000)
            initial_balances[b["id"]] = balance
            docs.append({
                "id": str(uuid.uuid4()),
                "bank_account_id": b["id"],
                "bank_account_name": b["name"],
                "snapshot_date": date_str,
                "balance": round(balance),
                "currency": "IDR",
                "source": "manual",
                "notes": f"Weekly snapshot {date_str}",
                "created_at": now_iso(),
                "updated_at": now_iso(),
            })
    if docs:
        await db.cash_balance_snapshots.insert_many(docs)
        for d in docs:
            d.pop("_id", None)
    log.info(f"[C.4] Inserted {len(docs)} cash snapshots")


# ============================================================
# C.5: Cash Accounts (for Cash Position dashboard)
# ============================================================
async def setup_cash_accounts(db):
    log.info("[C.5] Setting up cash accounts...")
    await db.cash_accounts.delete_many({})
    bank_accounts = []
    async for b in db.bank_accounts.find({}, {"_id": 0}):
        bank_accounts.append(b)
    docs = []
    for b in bank_accounts:
        docs.append({
            "id": str(uuid.uuid4()),
            "code": b["code"],
            "name": b["name"],
            "bank": b["bank"],
            "account_number": b["account_number"],
            "currency": "IDR",
            "current_balance": random.uniform(80_000_000, 500_000_000),
            "type": "bank",
            "active": True,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })
    # Add cash on hand
    docs.append({
        "id": str(uuid.uuid4()),
        "code": "CASH-OH",
        "name": "Cash on Hand",
        "bank": "Cash",
        "account_number": "-",
        "currency": "IDR",
        "current_balance": random.uniform(15_000_000, 50_000_000),
        "type": "cash",
        "active": True,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    })
    if docs:
        await db.cash_accounts.insert_many(docs)
        for d in docs:
            d.pop("_id", None)
    log.info(f"[C.5] Inserted {len(docs)} cash accounts")


# ============================================================
# C.6: Anomaly Events
# ============================================================
async def simulate_anomalies(db):
    log.info("[C.6] Simulating anomaly events...")
    await db.anomaly_events.delete_many({})

    outlets = []
    async for o in db.outlets.find({}, {"_id": 0}):
        outlets.append(o)
    vendors = []
    async for v in db.vendors.find({}, {"_id": 0}):
        vendors.append(v)

    types = ["sales_deviation", "vendor_price_spike", "vendor_leadtime", "ap_cash_spike"]
    docs = []
    for i in range(15):
        atype = random.choice(types)
        outlet = random.choice(outlets)
        severity = random.choices(["mild", "severe"], weights=[0.7, 0.3])[0]
        status = random.choices(["open", "investigating", "resolved", "false_positive"],
                                weights=[0.4, 0.2, 0.3, 0.1])[0]
        days_ago = random.randint(0, 30)
        detected_at = (datetime.now(timezone.utc) - timedelta(days=days_ago)).isoformat()
        deviation = random.uniform(15, 80)

        if atype == "sales_deviation":
            title = f"Sales Anomaly: {outlet['name']}"
            description = f"Sales drop {deviation:.1f}% vs 14-day baseline"
            stats = {"observed": 5_500_000, "baseline": 8_500_000, "z_score": -2.4, "deviation_pct": -deviation}
        elif atype == "vendor_price_spike":
            v = random.choice(vendors)
            title = f"Vendor Price Spike: {v['name']}"
            description = f"Item price up {deviation:.1f}% vs 90-day avg"
            stats = {"observed_price": 85000, "avg_price": 65000, "deviation_pct": deviation}
        elif atype == "vendor_leadtime":
            v = random.choice(vendors)
            title = f"Lead Time Delay: {v['name']}"
            description = f"Lead time +{int(deviation/10)} days vs baseline"
            stats = {"observed_days": 12, "baseline_days": 5, "delta_days": 7}
        else:
            title = f"AP Cash Spike: {outlet['name']}"
            description = f"Projected outflow {deviation:.1f}% above 3-month avg"
            stats = {"projected": 280_000_000, "avg_3mo": 200_000_000, "deviation_pct": deviation}

        docs.append({
            "id": str(uuid.uuid4()),
            "type": atype,
            "severity": severity,
            "status": status,
            "title": title,
            "description": description,
            "outlet_id": outlet["id"] if atype in ("sales_deviation", "ap_cash_spike") else None,
            "outlet_name": outlet["name"] if atype in ("sales_deviation", "ap_cash_spike") else None,
            "source_type": atype,
            "source_id": str(uuid.uuid4()),
            "stats": stats,
            "detected_at": detected_at,
            "resolved_at": now_iso() if status in ("resolved", "false_positive") else None,
            "triage_notes": [],
            "created_at": detected_at,
            "updated_at": detected_at,
        })
    if docs:
        await db.anomaly_events.insert_many(docs)
        for d in docs:
            d.pop("_id", None)
    log.info(f"[C.6] Inserted {len(docs)} anomaly events")


# ============================================================
# C.7: Notifications
# ============================================================
async def simulate_notifications(db):
    log.info("[C.7] Simulating notifications...")
    await db.notifications.delete_many({})

    user = await db.users.find_one({}, {"_id": 0})
    if not user:
        log.warning("[C.7] No user found, skipping notifications")
        return

    samples = [
        ("info", "Daily Sales Submitted", "Calluna All Day - Main submitted daily sales for today"),
        ("urgent", "Low Stock Alert", "12 items below par level at Altero Bistronomie"),
        ("warn", "Anomaly Detected", "Sales deviation >2σ at Maison de la Sol"),
        ("done", "Payment Approved", "PR-2604-0001 approved by CFO"),
        ("info", "GR Posted", "Goods received from PT Sumber Pangan"),
        ("warn", "AP Aging Warning", "5 invoices overdue >60 days"),
        ("info", "Period Closing Reminder", "Closing period 2026-04 due in 3 days"),
        ("urgent", "Cash Position Alert", "BCA 56 Torado below safety threshold"),
        ("done", "Forecast Updated", "Monthly sales forecast updated"),
        ("info", "New PO Created", "PO-2604-12345 sent to vendor"),
    ]
    docs = []
    for i in range(80):
        ntype, title, body = random.choice(samples)
        days_ago = random.randint(0, 30)
        created = (datetime.now(timezone.utc) - timedelta(days=days_ago, hours=random.randint(0, 23))).isoformat()
        is_read = random.random() < 0.6
        docs.append({
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "type": ntype,
            "title": title,
            "body": body,
            "link": None,
            "source_type": "system",
            "source_id": None,
            "read_at": now_iso() if is_read else None,
            "created_at": created,
            "updated_at": created,
        })
    if docs:
        await db.notifications.insert_many(docs)
        for d in docs:
            d.pop("_id", None)
    log.info(f"[C.7] Inserted {len(docs)} notifications")


# ============================================================
# C.8: Vouchers (already in Excel) + Tax Records
# ============================================================
async def import_vouchers_and_tax(db):
    log.info("[C.8] Importing vouchers + tax records from Excel...")
    import openpyxl
    EXCEL_FIN = "/app/excel_source/financial_report.xlsx"

    await db.vouchers.delete_many({})
    await db.tax_records.delete_many({})

    wb = openpyxl.load_workbook(EXCEL_FIN, data_only=True, read_only=True)

    # Vouchers
    def _to_float(v):
        try:
            return float(v) if v is not None and str(v).replace(".", "").replace(",", "").replace("-", "").isdigit() else 0
        except Exception:
            return 0

    if 'Voucher' in wb.sheetnames:
        ws = wb['Voucher']
        v_docs = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            code = str(row[1]) if row[1] else ""
            issued = row[2]
            expired = row[3]
            name = str(row[4]) if row[4] else ""
            value = _to_float(row[5]) if row[5] is not None else 0
            remarks = str(row[13]) if len(row) > 13 and row[13] else ""
            if not code or not name or "voucher code" in code.lower() or "name" in name.lower():
                continue
            v_docs.append({
                "id": str(uuid.uuid4()),
                "voucher_code": code,
                "issued_date": issued.strftime("%Y-%m-%d") if isinstance(issued, datetime) else None,
                "expired_date": expired.strftime("%Y-%m-%d") if isinstance(expired, datetime) else None,
                "recipient_name": name,
                "value": value,
                "claimed": False,
                "remarks": remarks,
                "created_at": now_iso(),
                "updated_at": now_iso(),
            })
        if v_docs:
            await db.vouchers.insert_many(v_docs)
            for d in v_docs:
                d.pop("_id", None)
        log.info(f"[C.8] Inserted {len(v_docs)} vouchers")

    # Tax records
    if 'Tax Details' in wb.sheetnames:
        ws = wb['Tax Details']
        t_docs = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            date = row[0]
            desc = str(row[1]) if row[1] else ""
            inv_no = str(row[2]) if row[2] else ""
            cr = _to_float(row[3])
            db_amt = _to_float(row[4])
            remaining = _to_float(row[8]) if len(row) > 8 else 0
            if not desc or "description" in desc.lower():
                continue
            t_docs.append({
                "id": str(uuid.uuid4()),
                "txn_date": date.strftime("%Y-%m-%d") if isinstance(date, datetime) else None,
                "description": desc,
                "invoice_no": inv_no,
                "credit": cr,
                "debit": abs(db_amt),
                "remaining": remaining,
                "tax_type": "PPN",
                "created_at": now_iso(),
                "updated_at": now_iso(),
            })
        if t_docs:
            await db.tax_records.insert_many(t_docs)
            for d in t_docs:
                d.pop("_id", None)
        log.info(f"[C.8] Inserted {len(t_docs)} tax records")
    wb.close()


# ============================================================
# C.9: Payment Requests (PR Phase 1 feature)
# ============================================================
async def simulate_payment_requests(db):
    log.info("[C.9] Simulating Payment Request workflow data...")
    await db.payment_requests.delete_many({})

    open_aps = []
    async for ap in db.ap_ledgers.find({"status": "open"}, {"_id": 0}).limit(500):
        open_aps.append(ap)

    # Group by week
    user = await db.users.find_one({}, {"_id": 0})
    user_id = user["id"] if user else "system"
    user_name = user.get("full_name", "Admin") if user else "System"

    docs = []
    seq = 0
    week_groups = defaultdict(list)
    for ap in open_aps:
        if ap.get("invoice_date"):
            try:
                d = datetime.fromisoformat(ap["invoice_date"].replace("Z", "+00:00")) if "T" in ap["invoice_date"] else datetime.strptime(ap["invoice_date"][:10], "%Y-%m-%d")
                week_key = f"{d.year}-W{d.isocalendar()[1]:02d}"
                week_groups[week_key].append(ap)
            except Exception:
                pass

    for week_key, aps in list(week_groups.items())[:25]:  # 25 weeks
        if not aps:
            continue
        # Pick subset
        selected = random.sample(aps, min(random.randint(2, 8), len(aps)))
        items = []
        total = 0
        for ap in selected:
            items.append({
                "ap_ledger_id": ap["id"],
                "vendor_id": ap["vendor_id"],
                "vendor_name": ap.get("vendor_name", ""),
                "invoice_no": ap.get("invoice_no", ""),
                "invoice_date": ap.get("invoice_date"),
                "due_date": ap.get("due_date"),
                "amount": ap["amount"],
                "priority": "normal",
            })
            total += ap["amount"]

        seq += 1
        request_date = aps[0].get("invoice_date") or now_iso()
        if isinstance(request_date, datetime):
            request_date = request_date.isoformat()
        status = random.choices(["draft", "submitted", "approved", "paid"], weights=[0.1, 0.2, 0.3, 0.4])[0]
        docs.append({
            "id": str(uuid.uuid4()),
            "doc_no": f"PR-{week_key}-{seq:03d}",
            "request_date": request_date[:10] if isinstance(request_date, str) else None,
            "period_week": week_key,
            "brand_id": None,
            "outlet_id": None,
            "items": items,
            "total_amount": total,
            "requested_by": user_id,
            "requested_by_name": user_name,
            "status": status,
            "approval_chain": [],
            "current_approver": None,
            "approved_by": user_id if status in ("approved", "paid") else None,
            "approved_at": now_iso() if status in ("approved", "paid") else None,
            "paid_at": now_iso() if status == "paid" else None,
            "notes": f"Pengajuan pembayaran mingguan {week_key}",
            "created_at": now_iso(),
            "updated_at": now_iso(),
            "created_by": user_id,
        })
    if docs:
        await db.payment_requests.insert_many(docs)
        for d in docs:
            d.pop("_id", None)
    log.info(f"[C.9] Inserted {len(docs)} payment requests")


# ============================================================
# Main
# ============================================================
async def main():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    log.info("=" * 70)
    log.info("PHASE C — SMART SIMULATION LAYER")
    log.info("=" * 70)

    await simulate_daily_sales(db)
    await simulate_petty_cash(db)
    await simulate_customers(db)
    await simulate_cash_snapshots(db)
    await setup_cash_accounts(db)
    await simulate_anomalies(db)
    await simulate_notifications(db)
    await import_vouchers_and_tax(db)
    await simulate_payment_requests(db)

    log.info("=" * 70)
    log.info("PHASE C COMPLETE")
    log.info("=" * 70)
    client.close()


if __name__ == "__main__":
    asyncio.run(main())
